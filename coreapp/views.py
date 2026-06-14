from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.http import require_POST
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse, Http404
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView
)
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from django.db.models import Sum, Count, Q
from django.utils import timezone
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import (
    RuasJalan, SegmenJalan, Kecelakaan, AnalisisZScore, RekapSegmen,
    Kota, Kecamatan, Kelurahan, ClusterData, AIConfig, Profile, Polres,
    KecelakaanRaw, KecelakaanPreprosesing, LakaMentah
)
from rest_framework import status
import json
import math
import numpy as np
import requests
from datetime import datetime
import os

from coreapp.models import Polres

import pandas as pd
from sklearn.cluster import KMeans
from django.conf import settings
from io import StringIO, BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.discriminant_analysis import StandardScaler as LdaScaler
from sklearn.cluster import AgglomerativeClustering
from sklearn.metrics import silhouette_score
from django.core.paginator import Paginator
from .forms import (
    LoginForm, AdminCreateForm, AdminUpdateForm, PolresForm,
    RuasJalanForm, SegmenJalanForm,
    KecelakaanForm, RekapSegmenForm, UploadKecelakaanRawForm, UploadKecelakaanPreprosesForm
)

User = get_user_model()



# ============================================================
# HELPER FUNCTIONS — Role Checking
# ============================================================
def is_admin(user):
    """Cek apakah user adalah admin atau superadmin (berdasarkan User.role)"""
    if not user.is_authenticated:
        return False
    try:
        return user.role in ('admin', 'superadmin') and user.is_active
    except Exception:
        return False


def is_superadmin(user):
    """Cek apakah user adalah superadmin"""
    if not user.is_authenticated:
        return False
    return user.role == 'superadmin' and user.is_active


def superadmin_required(view_func):
    return login_required(login_url='login')(
        user_passes_test(is_superadmin, login_url='dashboard')(view_func)
    )


# ============================================================
# AUTHENTICATION VIEWS
# ============================================================
def register_view(request):
    """Registrasi publik dinonaktifkan — hanya superadmin yang bisa membuat akun."""
    messages.error(request, 'Pendaftaran akun tidak tersedia. Hubungi Super Admin.')
    return redirect('login')


@login_required
def polres_list(request):
    if request.user.role != 'superadmin':
        return redirect('/')

    polres_list = Polres.objects.all()
    active_count = polres_list.filter(is_active=True).count()

    return render(request, 'coreapp/polres/list.html', {
        'polres_list': polres_list,
        'active_count': active_count
    })

@login_required
def polres_create(request):
    if request.user.role != 'superadmin':
        return redirect('/')

    form = PolresForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, 'Polres berhasil ditambahkan.')
        return redirect('polres_list')

    return render(request, 'coreapp/polres/form.html', {
        'form': form,
        'title': 'Tambah Polres Baru',
        'polres': None
    })

@login_required
def polres_update(request, pk):
    if request.user.role != 'superadmin':
        return redirect('/')

    polres = get_object_or_404(Polres, pk=pk)
    form = PolresForm(request.POST or None, instance=polres)
    if form.is_valid():
        form.save()
        messages.success(request, 'Polres berhasil diperbarui.')
        return redirect('polres_list')

    return render(request, 'coreapp/polres/form.html', {
        'form': form,
        'title': f'Edit Polres: {polres.nama}',
        'polres': polres
    })
@login_required
def polres_delete(request, pk):
    if request.user.role != 'superadmin':
        return redirect('/')

    polres = get_object_or_404(Polres, pk=pk)
    if request.method == 'POST':
        polres_name = polres.nama
        polres.delete()
        messages.success(request, f'Polres "{polres_name}" berhasil dihapus.')
        return redirect('polres_list')

    return render(
        request,
        'coreapp/polres/polres_confirm_delete.html',
        {'polres': polres}
    )
def login_view(request):
    """View login berbasis EMAIL + PASSWORD dengan validasi role dan is_active."""
    if request.user.is_authenticated:
        return redirect('dashboard')

    # Tangani pesan error dari Google OAuth redirect
    google_error = request.GET.get('error', '')
    google_error_messages = {
        'google_no_email': 'Akun Google tidak memiliki email yang dapat digunakan',
        'google_not_registered': 'Email Google Anda tidak terdaftar',
        'google_suc'
        'google_inactive': 'Akun Anda telah dinonaktifkan',
        'google_role_denied': 'Akun Anda tidak memiliki akses ke sistem ini',
        'google_no_profile': 'Profil akun tidak ditemukan',
    }
    if google_error and google_error in google_error_messages:
        messages.error(request, google_error_messages[google_error])

    form = LoginForm()

    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email'].strip().lower()
            password = form.cleaned_data['password']

            # 1. Cari user berdasarkan email
            try:
                user_obj = User.objects.get(email__iexact=email)
            except User.DoesNotExist:
                messages.error(request, 'Email Anda tidak terdaftar.')
                return render(request, 'registration/login.html', {'form': form})
            except User.MultipleObjectsReturned:
                user_obj = User.objects.filter(email__iexact=email).order_by('-created_at').first()
                if not user_obj:
                    messages.error(request, 'Email Anda tidak terdaftar.')
                    return render(request, 'registration/login.html', {'form': form})

            # 2. Cek is_active dari User model langsung
            if not user_obj.is_active:
                messages.error(request, 'Akun Anda telah dinonaktifkan.')
                return render(request, 'registration/login.html', {'form': form})

            # 3. Cek role harus superadmin atau admin
            if user_obj.role not in ('superadmin', 'admin'):
                messages.error(request, 'Akun Anda tidak memiliki akses ke sistem ini.')
                return render(request, 'registration/login.html', {'form': form})

            # 4. Autentikasi password via custom EmailBackend
            user = authenticate(request, username=email, password=password)
            if user is None:
                messages.error(request, 'Password salah. Silakan coba lagi.')
                return render(request, 'registration/login.html', {'form': form})

            # 5. Login berhasil
            login(request, user, backend='coreapp.backends.EmailBackend')
            nama = user.first_name or user.name or user.email
            messages.success(request, f'Selamat datang, {nama}!')
            next_url = request.GET.get('next', 'dashboard')
            return redirect(next_url)
        # else:
        #     messages.error(request, 'Data yang dimasukkan tidak valid.')

    return render(request, 'registration/login.html', {'form': form})



def logout_view(request):
    """View untuk logout"""
    logout(request)
    messages.success(request, 'Anda telah logout.')
    return redirect('login')


# ============================================================
# SUPERADMIN — KELOLA AKUN ADMIN
# ============================================================
@superadmin_required
def admin_list(request):
    """Daftar semua akun admin (hanya superadmin)"""
    # Ambil semua user yang punya profile dengan role admin/superadmin
    profiles = Profile.objects.select_related('user').filter(
        role__in=['admin', 'superadmin']
    ).order_by('role', 'user__email')

    context = {
        'profiles': profiles,
        'polres_choices': Polres.objects.all(),
    }
    return render(request, 'superadmin/admin_list.html', context)


@superadmin_required
def admin_create(request):
    """Buat akun admin baru (hanya superadmin)"""
    if request.method == 'POST':
        form = AdminCreateForm(request.POST)
        if form.is_valid():
            user = form.save()
            nama = f"{user.first_name} {user.last_name}".strip() or user.email
            messages.success(request, f'Akun admin "{nama}" berhasil dibuat.')
            return redirect('admin_list')
    else:
        form = AdminCreateForm()

    context = {
        'form': form,
        'title': 'Tambah Akun Admin',
        'action': 'create',
    }
    return render(request, 'superadmin/admin_form.html', context)


@superadmin_required
def admin_update(request, user_id):
    """Edit akun admin (hanya superadmin)"""
    target_user = get_object_or_404(User, pk=user_id)

    # Superadmin tidak bisa edit dirinya sendiri di sini (pakai profil)
    if target_user == request.user:
        messages.warning(request, 'Gunakan halaman Profil untuk mengubah akun Anda sendiri.')
        return redirect('admin_list')

    try:
        target_profile = target_user.profile
    except Profile.DoesNotExist:
        messages.error(request, 'Profil tidak ditemukan untuk user ini.')
        return redirect('admin_list')

    if request.method == 'POST':
        form = AdminUpdateForm(request.POST, user_instance=target_user)
        if form.is_valid():
            form.save(target_user)
            nama = f"{target_user.first_name} {target_user.last_name}".strip() or target_user.email
            messages.success(request, f'Akun "{nama}" berhasil diperbarui.')
            return redirect('admin_list')
    else:
        # Isi form dengan data user saat ini
        form = AdminUpdateForm(
            user_instance=target_user,
            initial={
                'email': target_user.email,
                'first_name': target_user.first_name,
                'last_name': target_user.last_name,
                'role': target_profile.role,
                'polres': target_profile.polres,
                'is_active': target_profile.is_active,
            }
        )

    context = {
        'form': form,
        'title': f'Edit Akun: {f"{target_user.first_name} {target_user.last_name}".strip() or target_user.email}',
        'target_user': target_user,
        'target_profile': target_profile,
        'action': 'update',
    }
    return render(request, 'superadmin/admin_form.html', context)


@superadmin_required
def admin_delete(request, user_id):
    """Hapus akun admin (hanya superadmin)"""
    target_user = get_object_or_404(User, pk=user_id)

    # Jangan hapus diri sendiri
    if target_user == request.user:
        messages.error(request, 'Anda tidak dapat menghapus akun Anda sendiri.')
        return redirect('admin_list')

    if request.method == 'POST':
        nama = f"{target_user.first_name} {target_user.last_name}".strip() or target_user.email
        target_user.delete()
        messages.success(request, f'Akun "{nama}" berhasil dihapus.')
        return redirect('admin_list')

    context = {
        'target_user': target_user,
        'target_profile': getattr(target_user, 'profile', None),
    }
    return render(request, 'superadmin/admin_confirm_delete.html', context)


@superadmin_required
def admin_toggle_active(request, user_id):
    """Toggle status aktif/nonaktif akun admin"""
    target_user = get_object_or_404(User, pk=user_id)

    if target_user == request.user:
        messages.error(request, 'Anda tidak dapat menonaktifkan akun Anda sendiri.')
        return redirect('admin_list')

    try:
        profile = target_user.profile
        profile.is_active = not profile.is_active
        profile.save()
        status_text = 'diaktifkan' if profile.is_active else 'dinonaktifkan'
        nama = f"{target_user.first_name} {target_user.last_name}".strip() or target_user.email
        messages.success(request, f'Akun "{nama}" berhasil {status_text}.')
    except Exception as e:
        messages.error(request, f'Gagal mengubah status akun: {e}')

    return redirect('admin_list')


# Homepage View
def homepage_view(request):
    """Halaman homepage untuk user biasa"""
    context = {
        'total_ruas': RuasJalan.objects.count(),
        'total_segmen': SegmenJalan.objects.count(),
        'total_kecelakaan': KecelakaanPreprosesing.objects.count(),
        'total_korban': KecelakaanPreprosesing.objects.aggregate(
            total=Sum('korban_meninggal') + Sum('korban_luka_berat') + Sum('korban_luka_ringan')
        )['total'] or 0,
    }
    return render(request, 'homepage.html', context)

# =====================================
# HALAMAN LANDING PAGE
# =====================================

def tentang_view(request):
    return render(request, 'tentang.html')

def keselamatan_view(request):
    return render(request, 'informasikeselamatan.html')

def fitur_view(request):
    return render(request, 'fitur.html')

def faq_view(request):
    return render(request, 'faq.html')

# Dashboard Views
@login_required(login_url='login')
def dashboard_view(request):
    """Dashboard utama - untuk admin/user yang login"""
    selected_polres = request.GET.get('polres')
    polres_instance = None
    if selected_polres and selected_polres.isdigit():
        polres_instance = get_object_or_404(Polres, id=int(selected_polres))
        
    ruas_qs = RuasJalan.objects.all()
    segmen_qs = SegmenJalan.objects.all()
    kecelakaan_qs = KecelakaanPreprosesing.objects.all()
    cluster_data = ClusterData.objects.all()
    
    if polres_instance:
        ruas_qs = ruas_qs.filter(polres=polres_instance)
        segmen_qs = segmen_qs.filter(ruas_jalan__polres=polres_instance)
        kecelakaan_qs = kecelakaan_qs.filter(polres=polres_instance)
        cluster_data = cluster_data.filter(polres=polres_instance)
        
    context = {
        'total_ruas': ruas_qs.count(),
        'total_segmen': segmen_qs.count(),
        'total_kecelakaan': kecelakaan_qs.count(),
        'total_korban': kecelakaan_qs.aggregate(
            total=Sum('korban_meninggal') + Sum('korban_luka_berat') + Sum('korban_luka_ringan')
        )['total'] or 0,
    }
    
    # Statistik tahun ini
    tahun_ini = timezone.now().year
    context['kecelakaan_tahun_ini'] = kecelakaan_qs.filter(
        tanggal__year=tahun_ini
    ).count()
    
    # Segmen dengan kecelakaan terbanyak
    context['top_segmen'] = segmen_qs.annotate(
        jumlah_kecelakaan=Count('kecelakaan_preprosesing')
    ).order_by('-jumlah_kecelakaan')[:5]

    # --- Statistik Data Cluster (Non-AHC Parameters) ---
    context['total_cluster_data'] = cluster_data.count()
    
    # 1. Distribusi Jenis Kendaraan (Semua data)
    context['top_vehicles'] = list(cluster_data.values('jenis_kendaraan').annotate(
        count=Count('id')
    ).order_by('-count'))
    
    # 2. Distribusi Hari (Top Days)
    context['day_dist'] = list(cluster_data.values('hari').annotate(
        count=Count('id')
    ).order_by('-count'))
    
    # 3. Top 5 TKP
    context['top_tkp'] = list(cluster_data.values('tkp').annotate(
        count=Count('id')
    ).order_by('-count')[:5])
    
    # 4. Top 5 Jam (Sesi Waktu)
    context['top_hours'] = list(cluster_data.values('jam').annotate(
        count=Count('id')
    ).order_by('-count')[:5])

    # 5. Distribusi Penyebab (Cleaned)
    all_cluster_data = list(cluster_data)
    causes_map = {}
    for item in all_cluster_data:
        cause = item.penyebab.strip().lower() if item.penyebab else 'tidak diketahui'
        causes_map[cause] = causes_map.get(cause, 0) + 1
        
    sorted_causes = [{'penyebab': k, 'count': v} for k, v in sorted(causes_map.items(), key=lambda x: x[1], reverse=True)]
    context['causes_dist'] = sorted_causes
    
    if sorted_causes:
        context['dominant_cause'] = sorted_causes[0]['penyebab']
        context['dominant_cause_count'] = sorted_causes[0]['count']
    else:
        context['dominant_cause'] = 'N/A'
        context['dominant_cause_count'] = 0
        
    # 6. Distribusi Umur (7 Groups)
    age_groups = {
        '< 15': 0,
        '15-24': 0,
        '25-34': 0,
        '35-44': 0,
        '45-54': 0,
        '55-64': 0,
        '65+': 0
    }
    
    for item in all_cluster_data:
        age = item.umur
        if age < 15:
            age_groups['< 15'] += 1
        elif 15 <= age <= 24:
            age_groups['15-24'] += 1
        elif 25 <= age <= 34:
            age_groups['25-34'] += 1
        elif 35 <= age <= 44:
            age_groups['35-44'] += 1
        elif 45 <= age <= 54:
            age_groups['45-54'] += 1
        elif 55 <= age <= 64:
            age_groups['55-64'] += 1
        else:
            age_groups['65+'] += 1
            
    context['age_dist'] = [{'range': k, 'count': v} for k, v in age_groups.items()]
    
    # Find top age range
    top_age_range = max(age_groups, key=age_groups.get) if age_groups else 'N/A'
    context['top_age_range'] = top_age_range
    context['top_age_range_count'] = age_groups.get(top_age_range, 0)

    # Sediakan daftar polres untuk dropdown filter
    context['polres_list'] = Polres.objects.all()
    context['selected_polres'] = int(selected_polres) if selected_polres and selected_polres.isdigit() else None

    return render(request, 'coreapp/dashboard.html', context)


# Ruas Jalan Views
@login_required(login_url='login')
def ruas_jalan_list(request):
    """Daftar ruas jalan"""
    from .models import Polres
    
    ruas_jalan = RuasJalan.objects.all()
    
    if request.GET.get('search'):
        search = request.GET.get('search')
        ruas_jalan = ruas_jalan.filter(
            Q(nama_ruas__icontains=search) |
            Q(wilayah__icontains=search)
        )
    
    # Filter berdasarkan polres
    selected_polres = request.GET.get('polres', '')
    if selected_polres and selected_polres != 'all':
        ruas_jalan = ruas_jalan.filter(polres=selected_polres)
    
    context = {
        'ruas_jalan': ruas_jalan,
        'is_admin': is_admin(request.user),
        'polres_choices': Polres.objects.all(),
        'selected_polres': selected_polres
    }
    return render(request, 'coreapp/ruas_jalan/list.html', context)


@login_required(login_url='login')
@user_passes_test(is_admin)
def ruas_jalan_create(request):
    """Buat ruas jalan baru"""
    if request.method == 'POST':
        form = RuasJalanForm(request.POST)
        if form.is_valid():
            ruas = form.save(commit=False)
            # Capture polres dari user.profile.polres
            ruas.polres = request.user.profile.polres
            ruas.save()
            # Auto-generate segmen jalan
            ruas.generate_segmen()
            messages.success(request, f'Ruas jalan "{ruas.nama_ruas}" berhasil ditambahkan dan segmen otomatis dibuat.')
            return redirect('ruas_jalan_list')
    else:
        form = RuasJalanForm()
    
    context = {'form': form, 'title': 'Tambah Ruas Jalan'}
    return render(request, 'coreapp/ruas_jalan/form.html', context)


@login_required(login_url='login')
def ruas_jalan_detail(request, pk):
    """Detail ruas jalan"""
    ruas = get_object_or_404(RuasJalan, pk=pk)
    segmen = ruas.segmen_jalan.all()
    
    context = {
        'ruas': ruas,
        'segmen': segmen,
        'is_admin': is_admin(request.user)
    }
    return render(request, 'coreapp/ruas_jalan/detail.html', context)


@login_required(login_url='login')
@user_passes_test(is_admin)
def ruas_jalan_update(request, pk):
    """Update ruas jalan"""
    ruas = get_object_or_404(RuasJalan, pk=pk)
    
    if request.method == 'POST':
        form = RuasJalanForm(request.POST, instance=ruas)
        if form.is_valid():
            ruas = form.save()
            # Auto-regenerate segmen jalan saat update
            ruas.generate_segmen()
            messages.success(request, f'Ruas jalan "{ruas.nama_ruas}" berhasil diperbarui dan segmen otomatis dibuat ulang.')
            return redirect('ruas_jalan_detail', pk=ruas.pk)
    else:
        form = RuasJalanForm(instance=ruas)
    
    context = {'form': form, 'ruas': ruas, 'title': 'Edit Ruas Jalan'}
    return render(request, 'coreapp/ruas_jalan/form.html', context)


@login_required(login_url='login')
@user_passes_test(is_admin)
def ruas_jalan_delete(request, pk):
    """Hapus ruas jalan"""
    ruas = get_object_or_404(RuasJalan, pk=pk)
    
    if request.method == 'POST':
        nama = ruas.nama_ruas
        ruas.delete()
        messages.success(request, f'Ruas jalan "{nama}" berhasil dihapus.')
        return redirect('ruas_jalan_list')
    
    context = {'ruas': ruas}
    return render(request, 'coreapp/ruas_jalan/confirm_delete.html', context)


@login_required(login_url='login')
@user_passes_test(is_admin)
def generate_segmen(request, pk):
    """Generate segmen otomatis"""
    ruas = get_object_or_404(RuasJalan, pk=pk)
    ruas.generate_segmen()
    messages.success(request, f'Segmen untuk "{ruas.nama_ruas}" berhasil dibuat.')
    return redirect('ruas_jalan_detail', pk=pk)


# Profil Views
@login_required
def profile_view(request):
    user = request.user

    if request.method == "POST":
        user.username = request.POST.get("username")
        user.email = request.POST.get("email")

        full_name = request.POST.get("full_name")
        if full_name:
            nama = full_name.split(" ", 1)
            user.first_name = nama[0]
            user.last_name = nama[1] if len(nama) > 1 else ""

        user.save()

        return redirect('profile')

    return render(request, 'profile.html')


# Kecelakaan Views
@login_required(login_url='login')
def kecelakaan_list(request):
    """Daftar kecelakaan"""
    kecelakaan = Kecelakaan.objects.all()
    
    if request.GET.get('search'):
        search = request.GET.get('search')
        kecelakaan = kecelakaan.filter(
            Q(desa__icontains=search) |
            Q(kecamatan__icontains=search) |
            Q(kabupaten_kota__icontains=search)
        )
    
    if request.GET.get('tahun'):
        tahun = request.GET.get('tahun')
        kecelakaan = kecelakaan.filter(tanggal__year=tahun)
    
    context = {
        'kecelakaan': kecelakaan[:100],  # Limit untuk performa
        'is_admin': is_admin(request.user),
        'tahun_options': range(2020, timezone.now().year + 1)
    }
    return render(request, 'coreapp/kecelakaan/list.html', context)


@login_required(login_url='login')
@user_passes_test(is_admin)
def kecelakaan_create(request):
    """Tambah data kecelakaan"""
    if request.method == 'POST':
        form = KecelakaanForm(request.POST)
        if form.is_valid():
            kecelakaan = form.save()
            messages.success(request, 'Data kecelakaan berhasil ditambahkan.')
            return redirect('kecelakaan_list')
    else:
        form = KecelakaanForm()
    
    context = {'form': form, 'title': 'Tambah Kecelakaan'}
    return render(request, 'coreapp/kecelakaan/form.html', context)


@login_required(login_url='login')
def kecelakaan_detail(request, pk):
    """Detail kecelakaan"""
    kecelakaan = get_object_or_404(Kecelakaan, pk=pk)
    
    context = {
        'kecelakaan': kecelakaan,
        'is_admin': is_admin(request.user)
    }
    return render(request, 'coreapp/kecelakaan/detail.html', context)
@login_required(login_url='login')
@user_passes_test(is_admin)
def kecelakaan_update(request, pk):
    """Update kecelakaan"""
    kecelakaan = get_object_or_404(Kecelakaan, pk=pk)
    
    if request.method == 'POST':
        form = KecelakaanForm(request.POST, instance=kecelakaan)
        if form.is_valid():
            kecelakaan = form.save()
            messages.success(request, 'Data kecelakaan berhasil diperbarui.')
            return redirect('kecelakaan_detail', pk=kecelakaan.pk)
    else:
        form = KecelakaanForm(instance=kecelakaan)
    
    context = {'form': form, 'kecelakaan': kecelakaan, 'title': 'Edit Kecelakaan'}
    return render(request, 'coreapp/kecelakaan/form.html', context)


@login_required(login_url='login')
@user_passes_test(is_admin)
def kecelakaan_delete(request, pk):
    """Hapus kecelakaan"""
    kecelakaan = get_object_or_404(Kecelakaan, pk=pk)
    
    if request.method == 'POST':
        kecelakaan.delete()
        messages.success(request, 'Data kecelakaan berhasil dihapus.')
        return redirect('kecelakaan_list')
    
    context = {'kecelakaan': kecelakaan, 'cancel_url': 'kecelakaan_list'}
    return render(request, 'coreapp/kecelakaan/confirm_delete.html', context)

# Map Views
@login_required(login_url='login')
def map_view(request):
    """Tampilkan peta interaktif dengan filter tahun & ruas"""

    tahun_param = request.GET.get('tahun')

    # 👉 Jika kosong = SEMUA TAHUN
    if tahun_param:
        try:
            tahun = int(tahun_param)
        except (ValueError, TypeError):
            tahun = 0
    else:
        tahun = 0  # Semua tahun

    # Hitung Z-Score
    try:
        if tahun == 0:
            # 🔥 Hitung Z-Score dari SEMUA DATA
            AnalisisZScore.calculate_zscore_all_years()
        else:
            if not AnalisisZScore.objects.filter(tahun=tahun).exists():
                AnalisisZScore.calculate_zscore(tahun)
    except Exception as e:
        print(f"Warning: Z-Score calculation failed: {e}")

    context = {
        'tahun': tahun,
        'tahun_options': range(2020, timezone.now().year + 1),
    }

    return render(request, 'coreapp/map/map.html', context)


def peta_user_view(request):
    """Tampilkan peta interaktif untuk user biasa (tanpa sidebar, standalone)"""
    tahun_param = request.GET.get('tahun')
    
    if tahun_param:
        try:
            tahun = int(tahun_param)
        except (ValueError, TypeError):
            tahun = 0
    else:
        tahun = 0  # Semua tahun
    
    # Hitung Z-Score jika belum ada
    try:
        if not AnalisisZScore.objects.filter(tahun=tahun).exists():
            AnalisisZScore.calculate_zscore(tahun)
    except Exception as e:
        print(f"Warning: Could not calculate Z-Score for {tahun}: {e}")
    
    context = {
        'tahun': tahun,
        'tahun_options': range(2020, timezone.now().year + 1)
    }
    return render(request, 'peta_user.html', context)


# API Views
@api_view(['GET'])
def api_segmen_geojson(request):
    """API untuk mendapatkan GeoJSON segmen jalan dengan Z-Score atau default blue untuk no accidents"""
    tahun_raw = request.GET.get('tahun')
    if not tahun_raw or tahun_raw == 'None' or tahun_raw == '0':
        tahun = 0
    else:
        try:
            tahun = int(tahun_raw)
        except (ValueError, TypeError):
            tahun = 0
    
    print(f"\n{'='*80}")
    print(f"📍 API: api_segmen_geojson called for tahun={tahun}")
    print(f"{'='*80}")
    
    # Ensure Z-Score calculation exists for this year
    if not AnalisisZScore.objects.filter(tahun=tahun).exists():
        try:
            AnalisisZScore.calculate_zscore(tahun)
            print(f"✓ Auto-calculated Z-Score for {tahun}")
        except Exception as e:
            print(f"⚠ Could not auto-calculate Z-Score: {e}")
    
    segmen_list = SegmenJalan.objects.select_related('ruas_jalan').all()
    print(f"📊 Found {segmen_list.count()} segments in database")
    
    features = []
    line_count = 0
    marker_count = 0
    
    for segmen in segmen_list:
        # Hitung jumlah kecelakaan di segmen ini untuk tahun tertentu
        if tahun == 0:
            accident_count = KecelakaanPreprosesing.objects.filter(
                segmen_jalan=segmen
            ).count()
        else:
            accident_count = KecelakaanPreprosesing.objects.filter(
                segmen_jalan=segmen,
                tanggal__year=tahun
            ).count()
        
        # PENTING: Jika tidak ada kecelakaan, selalu set AMAN (ignore Z-Score jika ada)
        if accident_count == 0:
            kategori = 'aman'
            zscore = -2.0
            color = '#1976d2'  # Blue
        else:
            # Ada kecelakaan - cari analisis Z-Score
            try:
                analisis = AnalisisZScore.objects.get(segmen_jalan=segmen, tahun=tahun)
                kategori = analisis.kategori
                zscore = float(analisis.nilai_zscore)
                color = analisis.get_kategori_display_color()
            except AnalisisZScore.DoesNotExist:
                # Ada kecelakaan tapi belum ada Z-Score → coba hitung
                try:
                    AnalisisZScore.calculate_zscore(tahun)
                    analisis = AnalisisZScore.objects.get(segmen_jalan=segmen, tahun=tahun)
                    kategori = analisis.kategori
                    zscore = float(analisis.nilai_zscore)
                    color = analisis.get_kategori_display_color()
                except Exception:
                    kategori = 'unknown'
                    zscore = 0
                    color = '#999999'
        
        # PENTING: Generate geometry dari lat/lon jika kosong
        geometry = None
        if segmen.geometry:
            try:
                parsed_geom = json.loads(segmen.geometry)
                # Convert MultiLineString to LineString untuk rendering yang lebih baik
                if parsed_geom.get('type') == 'MultiLineString':
                    coords = []
                    for line in parsed_geom.get('coordinates', []):
                        coords.extend(line)
                    geometry = {'type': 'LineString', 'coordinates': coords}
                    print(f"✓ Converted MultiLineString to LineString for segmen {segmen.id}")
                else:
                    geometry = parsed_geom
            except Exception as e:
                print(f"⚠ Error parsing geometry for segmen {segmen.id}: {e}")
                geometry = None
        
        # Fallback: Jika geometry kosong, buat dari lat/lon awal-akhir
        if not geometry:
            if segmen.lat_awal and segmen.lon_awal and segmen.lat_akhir and segmen.lon_akhir:
                geometry = {
                    'type': 'LineString',
                    'coordinates': [
                        [float(segmen.lon_awal), float(segmen.lat_awal)],
                        [float(segmen.lon_akhir), float(segmen.lat_akhir)]
                    ]
                }
                print(f"✓ Generated fallback geometry for segmen {segmen.id} from lat/lon")
            else:
                # Coba ambil dari ruas_jalan geometry
                if segmen.ruas_jalan.geometry:
                    try:
                        ruas_geom = json.loads(segmen.ruas_jalan.geometry)
                        if ruas_geom.get('type') == 'Feature':
                            ruas_geom = ruas_geom.get('geometry', {})
                        
                        if ruas_geom.get('type') == 'LineString':
                            geometry = ruas_geom
                        elif ruas_geom.get('type') == 'MultiLineString':
                            coords = []
                            for line in ruas_geom.get('coordinates', []):
                                coords.extend(line)
                            geometry = {'type': 'LineString', 'coordinates': coords}
                        
                        if geometry:
                            print(f"✓ Generated geometry for segmen {segmen.id} from ruas_jalan")
                    except Exception as e:
                        print(f"⚠ Error getting ruas geometry for segmen {segmen.id}: {e}")
        
        if geometry:
            # 1. Feature LineString (Garis Jalan)
            feature_line = {
                'type': 'Feature',
                'id': f"line_{segmen.id}",
                'properties': {
                    'type': 'line',
                    'segmen_id': segmen.id,
                    'ruas_id': segmen.ruas_jalan.id,
                    'ruas_nama': segmen.ruas_jalan.nama_ruas,
                    'km_awal': float(segmen.km_awal),
                    'km_akhir': float(segmen.km_akhir),
                    'panjang': float(segmen.panjang_segmen),
                    'kategori': kategori,
                    'zscore': zscore,
                    'color': color,
                    'accident_count': accident_count,
                    'nama_segmen': segmen.nama_segmen or f"Segmen {segmen.km_awal}-{segmen.km_akhir}",
                    'keterangan': segmen.keterangan or '',
                    'url': f'/kecelakaan/segmen/{segmen.id}/'
                },
                'geometry': geometry
            }
            features.append(feature_line)
            line_count += 1

            # 2. Feature Point - Marker AWAL segmen
            if geometry.get('coordinates') and len(geometry['coordinates']) > 0:
                coords = geometry['coordinates']
                start_point = coords[0]
                end_point = coords[-1]
                
                # Marker awal segmen
                feature_point_start = {
                    'type': 'Feature',
                    'id': f"point_start_{segmen.id}",
                    'properties': {
                        'type': 'segment_marker',
                        'marker_type': 'start',
                        'segmen_id': segmen.id,
                        'ruas_id': segmen.ruas_jalan.id,
                        'ruas_nama': segmen.ruas_jalan.nama_ruas,
                        'km_awal': float(segmen.km_awal),
                        'km_akhir': float(segmen.km_akhir),
                        'panjang': float(segmen.panjang_segmen),
                        'kategori': kategori,
                        'zscore': zscore,
                        'color': color,
                        'accident_count': accident_count,
                        'nama_segmen': segmen.nama_segmen or f"Segmen {segmen.km_awal}-{segmen.km_akhir}",
                        'keterangan': segmen.keterangan or '',
                        'url': f'/kecelakaan/segmen/{segmen.id}/'
                    },
                    'geometry': {
                        'type': 'Point',
                        'coordinates': start_point
                    }
                }
                features.append(feature_point_start)
                marker_count += 1
                
                # Marker akhir segmen
                feature_point_end = {
                    'type': 'Feature',
                    'id': f"point_end_{segmen.id}",
                    'properties': {
                        'type': 'segment_marker',
                        'marker_type': 'end',
                        'segmen_id': segmen.id,
                        'ruas_id': segmen.ruas_jalan.id,
                        'ruas_nama': segmen.ruas_jalan.nama_ruas,
                        'km_awal': float(segmen.km_awal),
                        'km_akhir': float(segmen.km_akhir),
                        'panjang': float(segmen.panjang_segmen),
                        'kategori': kategori,
                        'zscore': zscore,
                        'color': color,
                        'accident_count': accident_count,
                        'nama_segmen': segmen.nama_segmen or f"Segmen {segmen.km_awal}-{segmen.km_akhir}",
                        'keterangan': segmen.keterangan or '',
                        'url': f'/kecelakaan/segmen/{segmen.id}/'
                    },
                    'geometry': {
                        'type': 'Point',
                        'coordinates': end_point
                    }
                }
                features.append(feature_point_end)
                marker_count += 1
        else:
            print(f"❌ Segmen {segmen.id} ({segmen.ruas_jalan.nama_ruas}) - NO geometry available!")
    
    # Add markers untuk start/end dari setiap ruas jalan
    ruas_segments = {}
    for segmen in segmen_list:
        if segmen.ruas_jalan.id not in ruas_segments:
            ruas_segments[segmen.ruas_jalan.id] = []
        ruas_segments[segmen.ruas_jalan.id].append(segmen)
    
    # Create markers untuk ruas jalan start/end
    for ruas_id, segments in ruas_segments.items():
        # Sort by km_awal to find first and last segment
        sorted_segments = sorted(segments, key=lambda s: float(s.km_awal))
        if sorted_segments:
            ruas_jalan = sorted_segments[0].ruas_jalan
            
            # Marker untuk AWAL ruas jalan
            first_segmen = sorted_segments[0]
            if first_segmen.geometry:
                try:
                    geom = json.loads(first_segmen.geometry)
                    if geom.get('type') == 'LineString' and geom.get('coordinates'):
                        start_coord = geom['coordinates'][0]
                        feature_ruas_start = {
                            'type': 'Feature',
                            'id': f"ruas_start_{ruas_id}",
                            'properties': {
                                'type': 'segment_marker',
                                'marker_type': 'ruas_start',
                                'ruas_id': ruas_id,
                                'ruas_nama': ruas_jalan.nama_ruas,
                                'label': f'Awal Ruas: {ruas_jalan.nama_ruas}'
                            },
                            'geometry': {
                                'type': 'Point',
                                'coordinates': start_coord
                            }
                        }
                        features.append(feature_ruas_start)
                        marker_count += 1
                except Exception as e:
                    print(f"⚠ Error creating ruas start marker for {ruas_jalan.nama_ruas}: {e}")
            
            # Marker untuk AKHIR ruas jalan
            last_segmen = sorted_segments[-1]
            if last_segmen.geometry:
                try:
                    geom = json.loads(last_segmen.geometry)
                    if geom.get('type') == 'LineString' and geom.get('coordinates'):
                        end_coord = geom['coordinates'][-1]
                        feature_ruas_end = {
                            'type': 'Feature',
                            'id': f"ruas_end_{ruas_id}",
                            'properties': {
                                'type': 'segment_marker',
                                'marker_type': 'ruas_end',
                                'ruas_id': ruas_id,
                                'ruas_nama': ruas_jalan.nama_ruas,
                                'label': f'Akhir Ruas: {ruas_jalan.nama_ruas}'
                            },
                            'geometry': {
                                'type': 'Point',
                                'coordinates': end_coord
                            }
                        }
                        features.append(feature_ruas_end)
                        marker_count += 1
                except Exception as e:
                    print(f"⚠ Error creating ruas end marker for {ruas_jalan.nama_ruas}: {e}")
    
    geojson = {
        'type': 'FeatureCollection',
        'features': features
    }
    
    print(f"✅ Response: {line_count} lines, {marker_count} markers - {len(features)} total features")
    print(f"{'='*80}\n")
    
    return Response(geojson)


@api_view(['GET'])
def api_threshold_data(request):
    """API untuk mendapatkan threshold data per ruas jalan untuk dynamic legend"""
    from django.db.models import Avg, StdDev
    
    tahun_raw = request.GET.get('tahun')
    if not tahun_raw or tahun_raw == 'None' or tahun_raw == '0':
        tahun = 0
    else:
        try:
            tahun = int(tahun_raw)
        except (ValueError, TypeError):
            tahun = 0
    
    # Ensure Z-Score calculation exists
    if not AnalisisZScore.objects.filter(tahun=tahun).exists():
        try:
            AnalisisZScore.calculate_zscore(tahun)
        except Exception as e:
            print(f"Warning: Could not auto-calculate Z-Score: {e}")
    
    ruas_jalan_list = RuasJalan.objects.all().distinct()
    threshold_data = {}
    
    for ruas_jalan in ruas_jalan_list:
        segments_in_ruas = SegmenJalan.objects.filter(ruas_jalan=ruas_jalan)
        
        # Get all z-scores for this ruas jalan
        analisis_list = AnalisisZScore.objects.filter(
            tahun=tahun,
            segmen_jalan__in=segments_in_ruas
        )
        
        if analisis_list.exists():
            zscore_values = [float(a.nilai_zscore) for a in analisis_list]
            z_max = max(zscore_values)
            z_min = min(zscore_values)
            
            # Calculate interval
            if z_max != z_min:
                interval = (z_max - z_min) / 5
            else:
                interval = 1
            
            # Calculate thresholds
            t1 = z_min + (1 * interval)
            t2 = z_min + (2 * interval)
            t3 = z_min + (3 * interval)
            t4 = z_min + (4 * interval)
            
            # Count segments per kategori
            kategori_counts = {
                'sangat_tinggi': analisis_list.filter(kategori='sangat_tinggi').count(),
                'tinggi': analisis_list.filter(kategori='tinggi').count(),
                'sedang': analisis_list.filter(kategori='sedang').count(),
                'rendah': analisis_list.filter(kategori='rendah').count(),
                'sangat_rendah': analisis_list.filter(kategori='sangat_rendah').count(),
            }
            
            threshold_data[ruas_jalan.id] = {
                'nama': ruas_jalan.nama_ruas,
                'has_data': True,
                'z_max': round(z_max, 3),
                'z_min': round(z_min, 3),
                'interval': round(interval, 3),
                't4': round(t4, 3),  # sangat_tinggi threshold
                't3': round(t3, 3),  # tinggi threshold
                't2': round(t2, 3),  # sedang threshold
                't1': round(t1, 3),  # rendah threshold
                'kategori_counts': kategori_counts,
                'total_segments': segments_in_ruas.count(),
            }
        else:
            threshold_data[ruas_jalan.id] = {
                'nama': ruas_jalan.nama_ruas,
                'has_data': False,
                'z_max': 0.0,
                'z_min': 0.0,
                'interval': 0.0,
                't4': 0.0,
                't3': 0.0,
                't2': 0.0,
                't1': 0.0,
                'kategori_counts': {
                    'sangat_tinggi': 0,
                    'tinggi': 0,
                    'sedang': 0,
                    'rendah': 0,
                    'sangat_rendah': 0,
                },
                'total_segments': segments_in_ruas.count(),
            }
    
    return Response({
        'tahun': tahun,
        'ruas_data': threshold_data
    })


@api_view(['GET'])
@login_required(login_url='login')
def api_kecelakaan_geojson(request):
    """API untuk mendapatkan GeoJSON kecelakaan"""
    tahun = request.GET.get('tahun', timezone.now().year)
    
    kecelakaan = KecelakaanPreprosesing.objects.filter(
        tanggal__year=tahun,
        latitude__isnull=False,
        longitude__isnull=False
    )
    
    features = []
    for k in kecelakaan:
        feature = {
            'type': 'Feature',
            'id': k.id,
            'properties': {
                'kecelakaan_id': k.id,
                'tanggal': k.tanggal.isoformat(),
                'waktu': k.waktu.isoformat(),
                'lokasi': f"{k.desa}, {k.kecamatan}",
                'korban_meninggal': k.korban_meninggal,
                'korban_luka_berat': k.korban_luka_berat,
                'korban_luka_ringan': k.korban_luka_ringan,
                'total_korban': k.total_korban,
                'kerugian': float(k.kerugian_materi),
                'url': f'/kecelakaan/{k.id}/'
            },
            'geometry': {
                'type': 'Point',
                'coordinates': [float(k.longitude), float(k.latitude)]
            }
        }
        features.append(feature)
    
    geojson = {
        'type': 'FeatureCollection',
        'features': features
    }
    
    return Response(geojson)


@api_view(['GET'])
@login_required(login_url='login')
def api_geoapify_routing(request):
    """
    API untuk mendapatkan rute jalan dari Geoapify.
    Input: lat_awal, lon_awal, lat_akhir, lon_akhir
    Output: JSON berisi distance_km dan geometry
    """
    lat_awal = request.GET.get('lat_awal')
    lon_awal = request.GET.get('lon_awal')
    lat_akhir = request.GET.get('lat_akhir')
    lon_akhir = request.GET.get('lon_akhir')

    if not all([lat_awal, lon_awal, lat_akhir, lon_akhir]):
        return JsonResponse({'status': 'error', 'message': 'Parameter tidak lengkap'}, status=400)

    api_key = os.getenv('GEOAPIFY_API_KEY')
    if not api_key:
        return JsonResponse({'status': 'error', 'message': 'API Key Geoapify tidak ditemukan di .env'}, status=500)

    url = f"https://api.geoapify.com/v1/routing?waypoints={lat_awal},{lon_awal}|{lat_akhir},{lon_akhir}&mode=drive&apiKey={api_key}"

    try:
        response = requests.get(url, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            if 'features' in data and len(data['features']) > 0:
                route_feature = data['features'][0]
                distance_meters = route_feature['properties']['distance']
                distance_km = round(distance_meters / 1000, 3)
                
                return JsonResponse({
                    'status': 'success',
                    'distance_km': distance_km,
                    'geometry': route_feature['geometry']
                })
            else:
                return JsonResponse({'status': 'error', 'message': 'Rute tidak ditemukan'}, status=404)
        else:
            return JsonResponse({'status': 'error', 'message': f'API Error: {response.status_code}'}, status=response.status_code)
            
    except requests.exceptions.Timeout:
        return JsonResponse({'status': 'error', 'message': 'API Request timeout'}, status=504)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@api_view(['GET'])
@login_required(login_url='login')
def api_geoapify_reverse_geocoding(request):
    """
    API untuk mendapatkan informasi alamat dari koordinat (Geoapify).
    Input: lat, lon
    Output: JSON berisi nama_jalan, wilayah (kota/kab, provinsi)
    """
    lat = request.GET.get('lat')
    lon = request.GET.get('lon')

    if not all([lat, lon]):
        return JsonResponse({'status': 'error', 'message': 'Parameter tidak lengkap'}, status=400)

    api_key = os.getenv('GEOAPIFY_API_KEY')
    if not api_key:
        return JsonResponse({'status': 'error', 'message': 'API Key Geoapify tidak ditemukan di .env'}, status=500)

    url = f"https://api.geoapify.com/v1/geocode/reverse?lat={lat}&lon={lon}&apiKey={api_key}"

    try:
        response = requests.get(url, timeout=10)
        
        if response.status_code == 200:
            data = response.json()
            if 'features' in data and len(data['features']) > 0:
                properties = data['features'][0]['properties']
                
                # Ekstrak informasi yang dibutuhkan
                nama_jalan = properties.get('street', properties.get('name', 'Tidak diketahui'))
                kota = properties.get('city', properties.get('county', ''))
                provinsi = properties.get('state', '')
                
                wilayah = f"{kota}, {provinsi}".strip(", ")
                
                return JsonResponse({
                    'status': 'success',
                    'nama_jalan': nama_jalan,
                    'wilayah': wilayah,
                    'kota_kab': kota,
                    'provinsi': provinsi
                })
            else:
                return JsonResponse({'status': 'error', 'message': 'Data lokasi tidak ditemukan'}, status=404)
        else:
            return JsonResponse({'status': 'error', 'message': f'API Error: {response.status_code}'}, status=response.status_code)
            
    except requests.exceptions.Timeout:
        return JsonResponse({'status': 'error', 'message': 'API Request timeout'}, status=504)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def api_data_update_check(request):
    """
    API untuk mengecek apakah ada update data terbaru
    Mengembalikan timestamp dari last update KecelakaanPreprosesing
    Frontend bisa compare dengan last known timestamp untuk trigger refresh
    """
    from django.db.models import Max
    
    tahun = request.GET.get('tahun')
    if not tahun or tahun == 'None' or tahun == '0':
        tahun = 0
    else:
        try:
            tahun = int(tahun)
        except (ValueError, TypeError):
            tahun = 0
    
    try:
        # Get latest updated_at from KecelakaanPreprosesing for this year
        if tahun == 0:
            latest_kecelakaan = KecelakaanPreprosesing.objects.aggregate(
                latest_update=Max('updated_at'),
                latest_create=Max('created_at')
            )
        else:
            latest_kecelakaan = KecelakaanPreprosesing.objects.filter(
                tanggal__year=tahun
            ).aggregate(
                latest_update=Max('updated_at'),
                latest_create=Max('created_at')
            )
        
        # Get latest update timestamp
        latest_update = latest_kecelakaan['latest_update'] or latest_kecelakaan['latest_create']
        
        # Get latest AnalisisZScore calculation timestamp
        latest_zscore = AnalisisZScore.objects.filter(tahun=tahun).values('id').last()
        
        return JsonResponse({
            'status': 'success',
            'tahun': tahun,
            'last_data_update': latest_update.timestamp() if latest_update else None,
            'has_zscore': latest_zscore is not None
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


@api_view(['GET'])
@login_required(login_url='login')
def api_analisis_statistik(request):
    """API untuk mendapatkan statistik analisis"""

    tahun_param = request.GET.get('tahun')
    analisis = AnalisisZScore.objects.all()

    if tahun_param and tahun_param != 'None' and tahun_param != '0':
        try:
            tahun = int(tahun_param)
            analisis = analisis.filter(tahun=tahun)
        except ValueError:
            return Response(
                {'error': 'Parameter tahun harus berupa angka'},
                status=400
            )
    else:
        tahun = 0
        analisis = analisis.filter(tahun=0)

    statistik = {
        'sangat_tinggi': analisis.filter(kategori='sangat_tinggi').count(),
        'tinggi': analisis.filter(kategori='tinggi').count(),
        'sedang': analisis.filter(kategori='sedang').count(),
        'rendah': analisis.filter(kategori='rendah').count(),
        'sangat_rendah': analisis.filter(kategori='sangat_rendah').count(),
    }

    return Response({
        'tahun': tahun,
        'total_segmen': analisis.count(),
        'kategori': statistik
    })
# Analisis Views
@login_required(login_url='login')
def analisis_view(request):
    """Halaman analisis"""
    tahun = request.GET.get('tahun', timezone.now().year)
    
    # Hitung ulang analisis
    if request.method == 'POST' and is_admin(request.user):
        RekapSegmen.update_rekap(tahun)
        AnalisisZScore.calculate_zscore(tahun)
        messages.success(request, f'Analisis untuk tahun {tahun} berhasil dihitung.')
    
    # Ambil analisis
    analisis = AnalisisZScore.objects.filter(tahun=tahun).order_by('-nilai_zscore')
    
    # Statistik
    statistik = {
        'sangat_tinggi': analisis.filter(kategori='sangat_tinggi'),
        'tinggi': analisis.filter(kategori='tinggi'),
        'sedang': analisis.filter(kategori='sedang'),
        'rendah': analisis.filter(kategori='rendah'),
        'sangat_rendah': analisis.filter(kategori='sangat_rendah'),
    }
    
    context = {
        'analisis': analisis,
        'statistik': statistik,
        'tahun': tahun,
        'tahun_options': range(2020, timezone.now().year + 1),
        'is_admin': is_admin(request.user)
    }
    
    return render(request, 'coreapp/analisis/analisis.html', context)


@login_required(login_url='login')
def segmen_kecelakaan_detail(request, segmen_id):
    """Detail kecelakaan per segmen"""
    segmen = get_object_or_404(SegmenJalan, pk=segmen_id)
    
    tahun_raw = request.GET.get('tahun')
    if tahun_raw is None:
        # Default ketika awal redirect -> semua tahun (0)
        tahun = 0
    elif tahun_raw == '' or tahun_raw == '0' or tahun_raw == 'all':
        tahun = 0
    else:
        try:
            tahun = int(tahun_raw)
        except (ValueError, TypeError):
            tahun = 0
            
    if tahun == 0:
        kecelakaan = KecelakaanPreprosesing.objects.filter(
            segmen_jalan=segmen
        )
    else:
        kecelakaan = KecelakaanPreprosesing.objects.filter(
            segmen_jalan=segmen,
            tanggal__year=tahun
        )
    
    # Calculate totals for the summary section
    rekap = kecelakaan.aggregate(
        total_meninggal=Sum('korban_meninggal'),
        total_luka_berat=Sum('korban_luka_berat'),
        total_luka_ringan=Sum('korban_luka_ringan')
    )
    
    total_meninggal = rekap['total_meninggal'] or 0
    total_luka = (rekap['total_luka_berat'] or 0) + (rekap['total_luka_ringan'] or 0)
    
    try:
        analisis = AnalisisZScore.objects.get(segmen_jalan=segmen, tahun=tahun)
    except AnalisisZScore.DoesNotExist:
        analisis = None
    
    context = {
        'segmen': segmen,
        'kecelakaan': kecelakaan,
        'analisis': analisis,
        'tahun': tahun,
        'tahun_options': range(2020, timezone.now().year + 1),
        'total_meninggal': total_meninggal,
        'total_luka': total_luka,
        'is_admin': is_admin(request.user)
    }
    
    return render(request, 'coreapp/analisis/segmen_detail.html', context)


@login_required(login_url='login')
def kecelakaan_raw_detail(request, pk):
    """Detail kecelakaan raw"""
    kecelakaan = get_object_or_404(KecelakaanRaw, pk=pk)
    return render(request, 'coreapp/kecelakaan/detail.html', {'kecelakaan': kecelakaan, 'type': 'Raw'})


@login_required(login_url='login')
def kecelakaan_preprosesing_detail(request, pk):
    """Detail kecelakaan preprocessing"""
    kecelakaan = get_object_or_404(KecelakaanPreprosesing, pk=pk)
    return render(request, 'coreapp/kecelakaan/detail.html', {'kecelakaan': kecelakaan, 'type': 'Preprocessing'})


@login_required(login_url='login')
@user_passes_test(is_admin)
def kecelakaan_raw_delete(request, pk):
    """Hapus data kecelakaan raw"""
    kecelakaan = get_object_or_404(KecelakaanRaw, pk=pk)
    if request.method == 'POST':
        kecelakaan.delete()
        messages.success(request, 'Data kecelakaan raw berhasil dihapus.')
        return redirect('kecelakaan_raw_list')
    return render(request, 'coreapp/kecelakaan/confirm_delete.html', {'kecelakaan': kecelakaan, 'type': 'Raw', 'cancel_url': 'kecelakaan_raw_list'})


@login_required(login_url='login')
@user_passes_test(is_admin)
def kecelakaan_preprosesing_delete(request, pk):
    """Hapus data kecelakaan preprocessing"""
    kecelakaan = get_object_or_404(KecelakaanPreprosesing, pk=pk)
    if request.method == 'POST':
        kecelakaan.delete()
        messages.success(request, 'Data kecelakaan preprocessing berhasil dihapus.')
        return redirect('kecelakaan_preprosesing_list')
    return render(request, 'coreapp/kecelakaan/confirm_delete.html', {'kecelakaan': kecelakaan, 'type': 'Preprocessing', 'cancel_url': 'kecelakaan_preprosesing_list'})


# Cluster K-Means Views
@login_required(login_url='login')
def cluster_data(request):
    data = Kecelakaan.objects.all()[:50]

    context = {
        'kecelakaan': data
    }

    return render(request, 'coreapp/data_cluster/cluster.html', context)







# ===============================
# AHC VIEWS
# ===============================

# ================================
# HALAMAN DATA
# ================================

# =========================================================================================
# K-MEANS VIEWS (Imported from utils_kmeans.py)
# =========================================================================================
from .utils_kmeans import (
    preprocessing,
    reset_k_means,
    proses_cluster,
    hasil,
    rekomendasi_kebijakan,
    get_ai_recommendation,
    analyze_accident_clustering,
    save_ai_config,
    kmeans_data,
    kmeans_proses,
    kmeans_hasil
)


# ==========================================
# K-MEANS DATA MANAGEMENT
# ==========================================

from django.core.paginator import Paginator

def normalize_jam(jam_str):
    if not jam_str:
        return "00:00"
    jam_str = str(jam_str).strip().replace('.', ':')
    if ':' in jam_str:
        parts = jam_str.split(':')
        if len(parts) >= 2:
            h = parts[0].zfill(2)
            m = parts[1].ljust(2, '0')[:2]
            return f"{h}:{m}"
    if jam_str.isdigit():
        return f"{jam_str.zfill(2)}:00"
    return jam_str

@login_required(login_url='login')
def cluster_data_list(request):
    all_data = ClusterData.objects.all()
    
    # Filter Tahun
    selected_tahun = request.GET.get('tahun')
    if selected_tahun:
        all_data = all_data.filter(tahun=selected_tahun)
        
    # Filter Polres
    selected_polres = request.GET.get('polres')
    if selected_polres and selected_polres.isdigit():
        all_data = all_data.filter(polres_id=int(selected_polres))
        
    total_count = all_data.count()
    
    # Hitung duplikasi (data yang isinya persis sama di semua kolom utama)
    duplicate_groups = all_data.values(
        'no_referensi', 'umur', 'tkp', 'penyebab', 'hari', 'tanggal', 'jam', 
        'jenis_kendaraan', 'tipe_kendaraan', 'kerugian_material', 'tahun', 'polres'
    ).annotate(count=Count('id')).filter(count__gt=1)
    
    jumlah_duplikat = sum(group['count'] - 1 for group in duplicate_groups)

    # Ambil detail data duplikat untuk ditampilkan di modal
    duplicate_data_details = []
    for group in duplicate_groups:
        # Ambil satu contoh data untuk setiap grup duplikat
        example = all_data.filter(
            no_referensi=group['no_referensi'],
            umur=group['umur'], tkp=group['tkp'], penyebab=group['penyebab'],
            hari=group['hari'], tanggal=group['tanggal'], jam=group['jam'],
            jenis_kendaraan=group['jenis_kendaraan'], tipe_kendaraan=group['tipe_kendaraan'],
            kerugian_material=group['kerugian_material']
        ).first()
        if example:
            duplicate_data_details.append({
                'example': example,
                'count': group['count']
            })

    data_list = all_data.order_by('-tanggal', '-jam')
    
    # Normalize jam for display (existing data might have 19:0)
    for d in data_list:
        d.jam = normalize_jam(d.jam)

    # Paginasi 20 item per halaman
    paginator = Paginator(data_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Generate list tahun
    years_list = list(range(2027, 1944, -1))
    filter_years = list(ClusterData.objects.exclude(tahun__isnull=True).values_list('tahun', flat=True).distinct().order_by('-tahun'))
    
    selected_tahun_int = int(selected_tahun) if selected_tahun and selected_tahun.isdigit() else None
    selected_polres_int = int(selected_polres) if selected_polres and selected_polres.isdigit() else None
    
    context = {
        'data_list': page_obj,
        'total_data': total_count,
        'jumlah_duplikat': jumlah_duplikat,
        'duplicate_data_details': duplicate_data_details,
        'ai_config': AIConfig.objects.filter(tipe='kmeans').first(),
        'ai_config_ahc': AIConfig.objects.filter(tipe='ahc').first(),
        'years_list': years_list,
        'filter_years': filter_years,
        'selected_tahun': selected_tahun_int,
        'selected_polres': selected_polres_int,
        'polres_list': Polres.objects.all(),
    }
    return render(request, 'coreapp/data_cluster/list.html', context)

@login_required(login_url='login')
def cluster_data_tambah(request):
    if request.method == "POST":
        umur = request.POST.get('umur')
        tkp = request.POST.get('tkp')
        penyebab = request.POST.get('penyebab')
        tanggal_raw = request.POST.get('tanggal') # Format YYYY-MM-DD
        jam = request.POST.get('jam')
        jenis_kendaraan = request.POST.get('jenis_kendaraan')
        tipe_kendaraan = request.POST.get('tipe_kendaraan')
        kerugian_material = request.POST.get('kerugian_material')
        
        tahun_val = request.POST.get('tahun')
        tahun = int(tahun_val) if tahun_val and tahun_val.isdigit() else None
        polres_id = request.POST.get('polres')
        polres_instance = None
        if polres_id and polres_id.isdigit():
            polres_instance = get_object_or_404(Polres, id=int(polres_id))

        if kerugian_material:
            kerugian_material = kerugian_material.upper().replace('RP', '').replace('.', '').replace(',', '').replace(' ', '').strip()

        tanggal_obj = datetime.strptime(tanggal_raw, '%Y-%m-%d')
        hari_en = tanggal_obj.strftime('%A')
        hari_map = {
            'Monday': 'Senin', 'Tuesday': 'Selasa', 'Wednesday': 'Rabu',
            'Thursday': 'Kamis', 'Friday': 'Jumat', 'Saturday': 'Sabtu', 'Sunday': 'Minggu'
        }
        hari = hari_map.get(hari_en, hari_en)

        ClusterData.objects.create(
            umur=umur,
            tkp=str(tkp).strip(),
            penyebab=str(penyebab).strip(),
            hari=str(hari).strip(),
            tanggal=tanggal_raw,
            jam=normalize_jam(jam),
            jenis_kendaraan=str(jenis_kendaraan).strip(),
            tipe_kendaraan=str(tipe_kendaraan).strip(),
            kerugian_material=kerugian_material,
            tahun=tahun,
            polres=polres_instance
        )
        messages.success(request, "Data berhasil ditambahkan.")
        return redirect('cluster_data_list')

    # Dropdown data: Ambil unik, hilangkan yang kosong/null, urutkan
    def get_unique_choices(field):
        # Ambil dari DB
        raw_choices = ClusterData.objects.exclude(**{f"{field}__isnull": True}).exclude(**{f"{field}": ""}).values_list(field, flat=True).distinct()
        
        # Bersihkan whitespace dan pastikan unik di Python (case-insensitive check or normalization)
        cleaned = set()
        for c in raw_choices:
            if c:
                cleaned.add(str(c).strip())
        
        return sorted(list(cleaned))

    years_list = list(range(2027, 1944, -1))
    polres_list = Polres.objects.all()

    context = {
        'tkp_choices': get_unique_choices('tkp'),
        'penyebab_choices': get_unique_choices('penyebab'),
        'jenis_choices': get_unique_choices('jenis_kendaraan'),
        'tipe_choices': get_unique_choices('tipe_kendaraan'),
        'years_list': years_list,
        'polres_list': polres_list,
    }
    return render(request, 'coreapp/data_cluster/tambah.html', context)

@login_required(login_url='login')
def cluster_data_edit(request, pk):
    from django.shortcuts import get_object_or_404
    data = get_object_or_404(ClusterData, pk=pk)
    
    if request.method == "POST":
        umur = request.POST.get('umur')
        tkp = request.POST.get('tkp')
        penyebab = request.POST.get('penyebab')
        tanggal_raw = request.POST.get('tanggal')
        jam = request.POST.get('jam')
        jenis_kendaraan = request.POST.get('jenis_kendaraan')
        tipe_kendaraan = request.POST.get('tipe_kendaraan')
        kerugian_material = request.POST.get('kerugian_material')
        
        tahun_val = request.POST.get('tahun')
        data.tahun = int(tahun_val) if tahun_val and tahun_val.isdigit() else None
        polres_id = request.POST.get('polres')
        if polres_id and polres_id.isdigit():
            data.polres = get_object_or_404(Polres, id=int(polres_id))
        else:
            data.polres = None

        if kerugian_material:
            kerugian_material = kerugian_material.upper().replace('RP', '').replace('.', '').replace(',', '').replace(' ', '').strip()

        tanggal_obj = datetime.strptime(tanggal_raw, '%Y-%m-%d')
        hari_en = tanggal_obj.strftime('%A')
        hari_map = {
            'Monday': 'Senin', 'Tuesday': 'Selasa', 'Wednesday': 'Rabu',
            'Thursday': 'Kamis', 'Friday': 'Jumat', 'Saturday': 'Sabtu', 'Sunday': 'Minggu'
        }
        hari = hari_map.get(hari_en, hari_en)

        data.umur = umur
        data.tkp = str(tkp).strip()
        data.penyebab = str(penyebab).strip()
        data.hari = str(hari).strip()
        data.tanggal = tanggal_raw
        data.jam = normalize_jam(jam)
        data.jenis_kendaraan = str(jenis_kendaraan).strip()
        data.tipe_kendaraan = str(tipe_kendaraan).strip()
        data.kerugian_material = kerugian_material
        data.save()
        
        messages.success(request, "Data berhasil diupdate.")
        return redirect('cluster_data_list')

    # Dropdown data: Ambil unik, hilangkan yang kosong/null, urutkan
    def get_unique_choices(field):
        raw_choices = ClusterData.objects.exclude(**{f"{field}__isnull": True}).exclude(**{f"{field}": ""}).values_list(field, flat=True).distinct()
        cleaned = set()
        for c in raw_choices:
            if c:
                cleaned.add(str(c).strip())
        return sorted(list(cleaned))

    years_list = list(range(2027, 1944, -1))
    polres_list = Polres.objects.all()

    context = {
        'data': data,
        'tkp_choices': get_unique_choices('tkp'),
        'penyebab_choices': get_unique_choices('penyebab'),
        'jenis_choices': get_unique_choices('jenis_kendaraan'),
        'tipe_choices': get_unique_choices('tipe_kendaraan'),
        'years_list': years_list,
        'polres_list': polres_list,
    }
    return render(request, 'coreapp/data_cluster/edit.html', context)

def _parse_indo_date(date_str):
    """Helper to parse Indonesian date strings like '1 Januari 2024'"""
    if not date_str or pd.isna(date_str):
        return None
    
    if isinstance(date_str, (datetime, pd.Timestamp)):
        return date_str.date()
        
    date_str = str(date_str).strip()
    months_map = {
        'januari': 1, 'februari': 2, 'maret': 3, 'april': 4,
        'mei': 5, 'juni': 6, 'juli': 7, 'agustus': 8,
        'september': 9, 'oktober': 10, 'november': 11, 'desember': 12
    }
    
    try:
        # Split '1 Januari 2024'
        parts = date_str.split()
        if len(parts) == 3:
            day = int(parts[0])
            month_str = parts[1].lower()
            year = int(parts[2])
            month = months_map.get(month_str)
            if month:
                return datetime(year, month, day).date()
    except:
        pass
    
    # Try common formats using pandas
    try:
        return pd.to_datetime(date_str).date()
    except:
        return None

@login_required(login_url='login')
def cluster_data_import(request):
    if request.method == "POST":
        file = request.FILES.get('file')
        tahun_val = request.POST.get('tahun')
        tahun = int(tahun_val) if tahun_val and tahun_val.isdigit() else None
        polres_id = request.POST.get('polres')
        polres_instance = None
        if polres_id and polres_id.isdigit():
            polres_instance = get_object_or_404(Polres, id=int(polres_id))

        if not file:
            messages.error(request, "Silakan pilih file excel.")
            return redirect('cluster_data_list')
        
        try:
            df = pd.read_excel(file)
            df.columns = df.columns.str.strip()
            
            # Mapping kolom
            col_map = {}
            for col in df.columns:
                low = col.lower().replace(' ', '_')
                if 'jam' in low:               col_map[col] = 'jam'
                elif 'hari' in low:            col_map[col] = 'hari'
                elif 'tanggal' in low:         col_map[col] = 'tanggal'
                elif 'no' == low:              col_map[col] = 'no'
                elif 'umur' in low or 'usia' in low: col_map[col] = 'umur'
                elif 'tkp' in low or 'lokasi' in low: col_map[col] = 'tkp'
                elif 'penyebab' in low:        col_map[col] = 'penyebab'
                elif 'jenis_kendaraan' in low or 'jenis kendaraan' == col.lower(): col_map[col] = 'jenis_kendaraan'
                elif 'tipe_kendaraan' in low or 'tipe kendaraan' == col.lower():  col_map[col] = 'tipe_kendaraan'
                elif 'kerugian' in low:        col_map[col] = 'kerugian_material'
            
            df = df.rename(columns=col_map)
            
            count = 0
            hari_map_indo = {
                'Monday': 'Senin', 'Tuesday': 'Selasa', 'Wednesday': 'Rabu',
                'Thursday': 'Kamis', 'Friday': 'Jumat', 'Saturday': 'Sabtu', 'Sunday': 'Minggu'
            }

            for _, row in df.iterrows():
                try:
                    # Clean kerugian
                    kerugian = str(row.get('kerugian_material', 0))
                    kerugian = kerugian.upper().replace('RP', '').replace('.', '').replace(',', '').replace(' ', '').strip()
                    if kerugian == 'NAN' or not kerugian or kerugian == '-': kerugian = 0

                    # Handle tanggal
                    tgl = _parse_indo_date(row.get('tanggal'))
                    if not tgl:
                        continue # Skip invalid date rows
                    
                    # Ensure hari is correct
                    hari = row.get('hari')
                    if pd.isna(hari) or not hari:
                        hari_en = tgl.strftime('%A')
                        hari = hari_map_indo.get(hari_en, hari_en)
                    
                    # Ensure jam format
                    jam = str(row.get('jam', '')).replace('.', ':') # normalize 19.00 to 19:00 if needed, but the user example shows 19.00
                    if jam == 'nan': jam = "00:00"
                    
                    ClusterData.objects.create(
                        umur=row.get('umur', 0),
                        tkp=str(row.get('tkp', '')).strip(),
                        penyebab=str(row.get('penyebab', '')).strip(),
                        hari=str(hari).strip(),
                        tanggal=tgl,
                        jam=normalize_jam(row.get('jam', '')),
                        jenis_kendaraan=str(row.get('jenis_kendaraan', '')).strip(),
                        tipe_kendaraan=str(row.get('tipe_kendaraan', '')).strip(),
                        kerugian_material=kerugian,
                        tahun=tahun,
                        polres=polres_instance
                    )
                    count += 1
                except Exception as e:
                    import traceback
                    print(f"Error importing row: {e}")
                    traceback.print_exc()
            
            messages.success(request, f"Berhasil mengimport {count} data.")
        except Exception as e:
            messages.error(request, f"Gagal mengimport data: {str(e)}")
            
        return redirect('cluster_data_list')
    
    return redirect('cluster_data_list')

@login_required(login_url='login')
def cluster_data_hapus(request, pk):
    ClusterData.objects.filter(pk=pk).delete()
    messages.success(request, "Data berhasil dihapus.")
    return redirect('cluster_data_list')

@login_required(login_url='login')
def cluster_data_hapus_semua(request):
    from django.urls import reverse
    
    queryset = ClusterData.objects.all()
    
    # Filter Tahun
    tahun = request.GET.get('tahun')
    if tahun:
        queryset = queryset.filter(tahun=tahun)
        
    # Filter Polres
    polres_id = request.GET.get('polres')
    if polres_id and polres_id.isdigit():
        queryset = queryset.filter(polres_id=int(polres_id))
        
    count = queryset.count()
    queryset.delete()
    
    if tahun or polres_id:
        messages.success(request, f"Sebanyak {count} data Preprosesing berdasarkan filter yang dipilih berhasil dibersihkan.")
    else:
        messages.success(request, "Seluruh data Preprosesing berhasil dibersihkan.")
        
    url = reverse('cluster_data_list')
    query_params = []
    if tahun:
        query_params.append(f"tahun={tahun}")
    if polres_id:
        query_params.append(f"polres={polres_id}")
    if query_params:
        url += "?" + "&".join(query_params)
        
    return redirect(url)

@login_required(login_url='login')
def cluster_data_hapus_duplikat(request):
    if request.method == "POST":
        duplicate_groups = ClusterData.objects.values(
            'no_referensi', 'umur', 'tkp', 'penyebab', 'hari', 'tanggal', 'jam', 
            'jenis_kendaraan', 'tipe_kendaraan', 'kerugian_material'
        ).annotate(count=Count('id')).filter(count__gt=1)
        
        deleted_count = 0
        for group in duplicate_groups:
            ids = list(ClusterData.objects.filter(
                no_referensi=group['no_referensi'],
                umur=group['umur'], tkp=group['tkp'], penyebab=group['penyebab'],
                hari=group['hari'], tanggal=group['tanggal'], jam=group['jam'],
                jenis_kendaraan=group['jenis_kendaraan'], tipe_kendaraan=group['tipe_kendaraan'],
                kerugian_material=group['kerugian_material']
            ).values_list('id', flat=True))
            
            # Keep one, delete the rest
            ids_to_delete = ids[1:]
            ClusterData.objects.filter(id__in=ids_to_delete).delete()
            deleted_count += len(ids_to_delete)
            
        messages.success(request, f"Berhasil menghapus {deleted_count} data duplikat.")
        return redirect('cluster_data_list')
    return redirect('cluster_data_list')

# ==========================================
# END K-MEANS SECTION
# ==========================================

# View Tambah Data Kecelakaan
# =========================
@login_required(login_url='login')
def tambah_data(request):
    if request.method == "POST":
        # ambil data dari form
        tanggal = request.POST.get('tanggal')
        waktu = request.POST.get('waktu')
        meninggal = request.POST.get('meninggal') or 0
        luka_berat = request.POST.get('luka_berat') or 0
        luka_ringan = request.POST.get('luka_ringan') or 0
        kerugian = request.POST.get('kerugian') or 0
        kota = request.POST.get('kota')
        kecamatan = request.POST.get('kecamatan')
        kelurahan = request.POST.get('kelurahan')

        # simpan ke database
        Kecelakaan.objects.create(
            tanggal=tanggal,
            waktu=waktu,
            meninggal=meninggal,
            luka_berat=luka_berat,
            luka_ringan=luka_ringan,
            kerugian=kerugian,
            kota=kota,
            kecamatan=kecamatan,
            kelurahan=kelurahan
        )

        return redirect('data_cluster')  # kembali ke dashboard

    # Jika GET request → tampilkan halaman form
    return render(request, 'coreapp/data_cluster/tambah_kejadian.html')

def tambah_data_view(request):
    kota_list = Kota.objects.all()  # ambil semua kota
    return render(request, 'coreapp/data_cluster/tambah_kejadian.html', {
        'kota_list': kota_list
    })

def load_kecamatan(request):
    kota_id = request.GET.get('kota_id')
    kecamatan = list(Kecamatan.objects.filter(kota_id=kota_id).values('id', 'nama'))
    return JsonResponse({'kecamatan': kecamatan})

def load_kelurahan(request):
    kecamatan_id = request.GET.get('kecamatan_id')
    kelurahan = list(Kelurahan.objects.filter(kecamatan_id=kecamatan_id).values('id', 'nama'))
    return JsonResponse({'kelurahan': kelurahan})


# ================================


# ======================== Upload Kecelakaan Views ========================

@login_required(login_url='login')
@user_passes_test(is_admin)
def upload_kecelakaan_raw(request):
    """Upload data kecelakaan raw dari Excel/CSV"""
    if request.method == 'POST':
        form = UploadKecelakaanRawForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            try:
                # Parse Excel/CSV file
                if file.name.endswith('.csv'):
                    df = pd.read_csv(file, encoding='utf-8-sig')
                else:
                    df = pd.read_excel(file)
                
                # Validasi kolom yang diperlukan
                required_columns = ['tanggal', 'waktu', 'latitude', 'longitude', 
                                   'korban_meninggal', 'korban_luka_berat', 
                                   'korban_luka_ringan', 'kerugian_materi', 
                                   'desa', 'kecamatan', 'kabupaten_kota', 'keterangan']
                
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    messages.error(request, f'Kolom yang hilang: {", ".join(missing_columns)}')
                    return render(request, 'coreapp/kecelakaan/upload_raw.html', {'form': form})
                
                # Check apakah nomor_kecelakaan ada (opsional)
                has_nomor = 'nomor_kecelakaan' in df.columns
                
                # Import data
                count = 0
                errors = []
                for idx, row in df.iterrows():
                    try:
                        # Parse waktu - handle berbagai format
                        waktu_obj = None
                        if pd.notna(row['waktu']):
                            waktu_val = row['waktu']
                            # Jika sudah dalam format time object
                            if isinstance(waktu_val, type(pd.Timestamp.now().time())):
                                waktu_obj = waktu_val
                            # Jika string
                            elif isinstance(waktu_val, str):
                                try:
                                    waktu_obj = pd.to_datetime(waktu_val).time()
                                except:
                                    raise ValueError(f"Format waktu '{waktu_val}' tidak valid (gunakan HH:MM:SS)")
                            # Jika Timestamp/datetime
                            else:
                                try:
                                    waktu_obj = pd.to_datetime(waktu_val).time()
                                except:
                                    raise ValueError(f"Kolom waktu berisi tanggal bukan jam. Gunakan format HH:MM:SS")
                        
                        # Build create dict with nomor_kecelakaan jika available
                        create_data = {
                            'tanggal': pd.to_datetime(row['tanggal']),
                            'waktu': waktu_obj,
                            'latitude': float(row['latitude']),
                            'longitude': float(row['longitude']),
                            'korban_meninggal': int(row['korban_meninggal']) if pd.notna(row['korban_meninggal']) else 0,
                            'korban_luka_berat': int(row['korban_luka_berat']) if pd.notna(row['korban_luka_berat']) else 0,
                            'korban_luka_ringan': int(row['korban_luka_ringan']) if pd.notna(row['korban_luka_ringan']) else 0,
                            'kerugian_materi': float(row['kerugian_materi']) if pd.notna(row['kerugian_materi']) else 0,
                            'desa': str(row['desa']) if pd.notna(row['desa']) else '',
                            'kecamatan': str(row['kecamatan']) if pd.notna(row['kecamatan']) else '',
                            'kabupaten_kota': str(row['kabupaten_kota']) if pd.notna(row['kabupaten_kota']) else '',
                            'keterangan': str(row['keterangan']) if pd.notna(row['keterangan']) else '',
                            'polres': request.user.profile.polres
                        }
                        
                        # Tambah nomor_kecelakaan jika ada di file
                        if has_nomor and pd.notna(row['nomor_kecelakaan']):
                            create_data['nomor_kecelakaan'] = str(row['nomor_kecelakaan']).strip()
                        
                        KecelakaanRaw.objects.create(**create_data)
                        count += 1
                    except Exception as e:
                        errors.append(f"Baris {idx + 2}: {str(e)}")
                
                if errors and len(errors) <= 10:
                    messages.warning(request, f'Berhasil import {count} data, tapi ada beberapa error:\n' + '\n'.join(errors[:5]))
                else:
                    messages.success(request, f'Berhasil import {count} data kecelakaan raw.')
                
                return redirect('kecelakaan_raw_list')
            except Exception as e:
                messages.error(request, f'Error saat memproses file: {str(e)}')
    else:
        form = UploadKecelakaanRawForm()
    
    return render(request, 'coreapp/kecelakaan/upload_raw.html', {'form': form})


@login_required(login_url='login')
def kecelakaan_raw_list(request):
    """Daftar data kecelakaan raw"""
    from .models import Polres
    
    kecelakaan = KecelakaanRaw.objects.all()
    
    if request.GET.get('search'):
        search = request.GET.get('search')
        kecelakaan = kecelakaan.filter(
            Q(desa__icontains=search) |
            Q(kecamatan__icontains=search) |
            Q(kabupaten_kota__icontains=search)
        )
    
    if request.GET.get('tahun'):
        tahun = request.GET.get('tahun')
        kecelakaan = kecelakaan.filter(tanggal__year=tahun)
    
    # Filter berdasarkan polres
    selected_polres = request.GET.get('polres', '')
    if selected_polres and selected_polres != 'all':
        kecelakaan = kecelakaan.filter(polres=selected_polres)
    
    context = {
        'kecelakaan': kecelakaan[:100],
        'is_admin': is_admin(request.user),
        'tahun_options': range(2020, timezone.now().year + 1),
        'title': 'Data Kecelakaan Raw',
        'polres_choices': Polres.objects.all(),
        'selected_polres': selected_polres
    }
    return render(request, 'coreapp/kecelakaan/list_raw.html', context)


@login_required(login_url='login')
@user_passes_test(is_admin)
def upload_kecelakaan_preprosesing(request):
    """Upload data kecelakaan preprocessing dari Excel/CSV"""
    if request.method == 'POST':
        form = UploadKecelakaanPreprosesForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            try:
                # Parse Excel/CSV file
                if file.name.endswith('.csv'):
                    df = pd.read_csv(file, encoding='utf-8-sig')
                else:
                    df = pd.read_excel(file)
                
                # Validasi kolom yang diperlukan
                required_columns = ['tanggal', 'waktu', 'latitude', 'longitude', 
                                   'korban_meninggal', 'korban_luka_berat', 
                                   'korban_luka_ringan', 'kerugian_materi', 
                                   'desa', 'kecamatan', 'kabupaten_kota', 'keterangan']
                
                missing_columns = [col for col in required_columns if col not in df.columns]
                if missing_columns:
                    messages.error(request, f'Kolom yang hilang: {", ".join(missing_columns)}')
                    return render(request, 'coreapp/kecelakaan/upload_preprosesing.html', {'form': form})
                
                # Check apakah nomor_kecelakaan ada (opsional)
                has_nomor = 'nomor_kecelakaan' in df.columns
                
                # Import data
                count = 0
                errors = []
                for idx, row in df.iterrows():
                    try:
                        # Parse waktu - handle berbagai format
                        waktu_obj = None
                        if pd.notna(row['waktu']):
                            waktu_val = row['waktu']
                            # Jika sudah dalam format time object
                            if isinstance(waktu_val, type(pd.Timestamp.now().time())):
                                waktu_obj = waktu_val
                            # Jika string
                            elif isinstance(waktu_val, str):
                                try:
                                    waktu_obj = pd.to_datetime(waktu_val).time()
                                except:
                                    raise ValueError(f"Format waktu '{waktu_val}' tidak valid (gunakan HH:MM:SS)")
                            # Jika Timestamp/datetime
                            else:
                                try:
                                    waktu_obj = pd.to_datetime(waktu_val).time()
                                except:
                                    raise ValueError(f"Kolom waktu berisi tanggal bukan jam. Gunakan format HH:MM:SS")
                        
                        # Build create dict with nomor_kecelakaan jika available
                        create_data = {
                            'tanggal': pd.to_datetime(row['tanggal']),
                            'waktu': waktu_obj,
                            'latitude': float(row['latitude']),
                            'longitude': float(row['longitude']),
                            'korban_meninggal': int(row['korban_meninggal']) if pd.notna(row['korban_meninggal']) else 0,
                            'korban_luka_berat': int(row['korban_luka_berat']) if pd.notna(row['korban_luka_berat']) else 0,
                            'korban_luka_ringan': int(row['korban_luka_ringan']) if pd.notna(row['korban_luka_ringan']) else 0,
                            'kerugian_materi': float(row['kerugian_materi']) if pd.notna(row['kerugian_materi']) else 0,
                            'desa': str(row['desa']) if pd.notna(row['desa']) else '',
                            'kecamatan': str(row['kecamatan']) if pd.notna(row['kecamatan']) else '',
                            'kabupaten_kota': str(row['kabupaten_kota']) if pd.notna(row['kabupaten_kota']) else '',
                            'keterangan': str(row['keterangan']) if pd.notna(row['keterangan']) else '',
                            'polres': request.user.profile.polres
                        }
                        
                        # Tambah nomor_kecelakaan jika ada di file
                        if has_nomor and pd.notna(row['nomor_kecelakaan']):
                            create_data['nomor_kecelakaan'] = str(row['nomor_kecelakaan']).strip()
                        
                        KecelakaanPreprosesing.objects.create(**create_data)
                        count += 1
                    except Exception as e:
                        errors.append(f"Baris {idx + 2}: {str(e)}")
                
                if errors and len(errors) <= 10:
                    messages.warning(request, f'Berhasil import {count} data, tapi ada beberapa error:\n' + '\n'.join(errors[:5]))
                else:
                    messages.success(request, f'Berhasil import {count} data kecelakaan preprocessing.')
                
                return redirect('kecelakaan_preprosesing_list')
            except Exception as e:
                messages.error(request, f'Error saat memproses file: {str(e)}')
    else:
        form = UploadKecelakaanPreprosesForm()
    
    return render(request, 'coreapp/kecelakaan/upload_preprosesing.html', {'form': form})


@login_required(login_url='login')
def download_template_preprosesing(request):
    """Download template Excel untuk upload kecelakaan preprocessing"""
    # Buat workbook baru
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data Kecelakaan'
    
    # Definisikan kolom
    columns = [
        'nomor_kecelakaan',
        'tanggal',
        'waktu',
        'latitude',
        'longitude',
        'korban_meninggal',
        'korban_luka_berat',
        'korban_luka_ringan',
        'kerugian_materi',
        'desa',
        'kecamatan',
        'kabupaten_kota',
        'keterangan'
    ]
    
    # Styling untuk header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Tulis header
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Styling untuk data rows (example rows)
    data_alignment = Alignment(horizontal="left", vertical="center")
    
    # Tambahkan contoh baris kosong untuk format
    example_rows = 5
    for row_num in range(2, 2 + example_rows):
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = border
            cell.alignment = data_alignment
    
    # Set lebar kolom
    column_widths = {
        'A': 18,  # nomor_kecelakaan
        'B': 15,  # tanggal
        'C': 12,  # waktu
        'D': 12,  # latitude
        'E': 12,  # longitude
        'F': 18,  # korban_meninggal
        'G': 18,  # korban_luka_berat
        'H': 18,  # korban_luka_ringan
        'I': 16,  # kerugian_materi
        'J': 15,  # desa
        'K': 15,  # kecamatan
        'L': 18,  # kabupaten_kota
        'M': 20   # keterangan
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Set row height untuk header
    ws.row_dimensions[1].height = 25
    
    # Tambahkan sheet instruksi
    ws_instruction = wb.create_sheet('Instruksi')
    ws_instruction.column_dimensions['A'].width = 80
    
    instructions = [
        'INSTRUKSI PENGISIAN TEMPLATE DATA KECELAKAAN',
        '',
        'Kolom yang Wajib Diisi:',
        '• nomor_kecelakaan: Nomor urut atau ID kecelakaan (opsional, auto-generate jika kosong)',
        '• tanggal: Tanggal kejadian (format: YYYY-MM-DD, contoh: 2026-05-22)',
        '• waktu: Waktu kejadian (format: HH:MM:SS, contoh: 14:30:00)',
        '• latitude: Garis lintang (format: desimal, contoh: -6.2088)',
        '• longitude: Garis bujur (format: desimal, contoh: 106.8456)',
        '• korban_meninggal: Jumlah korban meninggal (angka)',
        '• korban_luka_berat: Jumlah korban luka berat (angka)',
        '• korban_luka_ringan: Jumlah korban luka ringan (angka)',
        '• kerugian_materi: Nilai kerugian materi (angka, dalam Rupiah)',
        '• desa: Nama desa/kelurahan (teks)',
        '• kecamatan: Nama kecamatan (teks)',
        '• kabupaten_kota: Nama kabupaten/kota (teks)',
        '• keterangan: Deskripsi atau catatan tambahan (teks)',
        '',
        'Catatan Penting:',
        '• Pastikan semua data sudah valid sebelum upload',
        '• Tanggal dan waktu harus sesuai format yang ditetapkan',
        '• Koordinat latitude/longitude harus dalam format desimal',
        '• Data akan otomatis diassign ke segmen jalan terdekat (threshold 5 km)',
        '• Maksimal 10.000 baris per file upload',
        '• Data tetap diimport ke sistem meskipun tidak cocok dengan segmen manapun',
    ]
    
    for row_num, instruction in enumerate(instructions, 1):
        cell = ws_instruction.cell(row=row_num, column=1)
        cell.value = instruction
        if row_num == 1:
            cell.font = Font(bold=True, size=12)
        elif instruction.startswith('•'):
            cell.alignment = Alignment(wrap_text=True)
    
    # Simpan ke BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return sebagai response
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Template_Kecelakaan_Preprosesing.xlsx"'
    
    return response


@login_required(login_url='login')
def download_template_clustering(request):
    """Download template Excel untuk upload data clustering (K-Means & AHC)"""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Data Clustering'
    
    # Definisikan kolom
    columns = [
        'No',
        'Hari',
        'Tanggal',
        'Jam',
        'Umur',
        'TKP',
        'Penyebab',
        'Jenis Kendaraan',
        'Tipe Kendaraan',
        'Kerugian Material'
    ]
    
    # Styling untuk header
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Tulis header
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        
    # Tambahkan contoh baris
    sample_data = [
        [1, 'Senin', '2026-05-25', '19:30', 25, 'Jl. Pahlawan', 'Mengantuk', 'Motor', 'Sepeda Motor', 1500000],
        [2, 'Rabu', '2026-05-20', '08:15', 42, 'Jl. Yos Sudarso', 'Rem Blong', 'Mobil', 'Minibus', 5000000],
        [3, 'Sabtu', '2026-05-23', '23:45', 19, 'Jl. Ahmad Yani', 'Kurang Konsentrasi', 'Motor', 'Sepeda Motor', 500000]
    ]
    
    data_alignment = Alignment(horizontal="left", vertical="center")
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = data_alignment
            
    # Set lebar kolom
    column_widths = {
        'A': 8,   # No
        'B': 12,  # Hari
        'C': 15,  # Tanggal
        'D': 12,  # Jam
        'E': 10,  # Umur
        'F': 25,  # TKP
        'G': 25,  # Penyebab
        'H': 18,  # Jenis Kendaraan
        'I': 18,  # Tipe Kendaraan
        'J': 20   # Kerugian Material
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
        
    # Set row height untuk header
    ws.row_dimensions[1].height = 25
    
    # Tambahkan sheet instruksi
    ws_instruction = wb.create_sheet('Instruksi')
    ws_instruction.column_dimensions['A'].width = 80
    
    instructions = [
        'INSTRUKSI PENGISIAN TEMPLATE DATA CLUSTERING (K-MEANS & AHC)',
        '',
        'Kolom yang Wajib Diisi:',
        '• No: Nomor urut data (angka)',
        '• Hari: Hari kejadian kecelakaan (contoh: Senin, Selasa, dst.)',
        '• Tanggal: Tanggal kejadian (format: YYYY-MM-DD, contoh: 2026-05-25)',
        '• Jam: Waktu kejadian (format: HH:MM atau HH.MM, contoh: 19:30 atau 19.30)',
        '• Umur: Umur korban kecelakaan (angka)',
        '• TKP: Tempat Kejadian Perkara / Lokasi (teks)',
        '• Penyebab: Faktor penyebab kecelakaan (contoh: Mengantuk, Rem Blong, Jalan Licin, Hujan, dll.)',
        '• Jenis Kendaraan: Jenis kendaraan (contoh: Motor, Mobil, Truk, Bus)',
        '• Tipe Kendaraan: Detail tipe kendaraan (contoh: Sepeda Motor, Minibus, Microbus, dll.)',
        '• Kerugian Material: Estimasi kerugian (angka desimal / bulat, contoh: 1500000)',
        '',
        'Catatan Penting:',
        '• Pastikan format Tanggal mengikuti YYYY-MM-DD agar proses pembacaan tanggal valid.',
        '• Kolom Penyebab akan otomatis dikategorikan oleh sistem ke dalam 4 faktor utama (Pengemudi, Jalan, Kendaraan, Lingkungan).',
        '• Kolom Jenis Kendaraan akan otomatis dikelompokkan ke dalam kategori (Motor, Mobil, Truk/Bus, Lainnya) untuk proses clustering.',
    ]
    
    for r_num, instruction in enumerate(instructions, 1):
        cell = ws_instruction.cell(row=r_num, column=1)
        cell.value = instruction
        if r_num == 1:
            cell.font = Font(bold=True, size=12)
        elif instruction.startswith('•'):
            cell.alignment = Alignment(wrap_text=True)
            
    # Simpan ke BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Template_Data_Clustering.xlsx"'
    return response


@login_required(login_url='login')
def kecelakaan_preprosesing_list(request):
    """Daftar data kecelakaan preprocessing"""
    from .models import Polres
    
    kecelakaan = KecelakaanPreprosesing.objects.all()
    
    if request.GET.get('search'):
        search = request.GET.get('search')
        kecelakaan = kecelakaan.filter(
            Q(nomor_kecelakaan__exact=search)
        )
    
    if request.GET.get('tahun'):
        tahun = request.GET.get('tahun')
        kecelakaan = kecelakaan.filter(tanggal__year=tahun)
    
    # Filter berdasarkan polres
    selected_polres = request.GET.get('polres', '')
    if selected_polres and selected_polres != 'all':
        kecelakaan = kecelakaan.filter(polres=selected_polres)
    
    context = {
        'kecelakaan': kecelakaan[:100],
        'is_admin': is_admin(request.user),
        'tahun_options': range(2020, timezone.now().year + 1),
        'title': 'Data Kecelakaan Preprosesing',
        'polres_choices': Polres.objects.all(),
        'selected_polres': selected_polres
    }
    return render(request, 'coreapp/kecelakaan/list_preprosesing.html', context)

# ===============================
# AHC VIEWS (Imported from utils_ahc.py)
# ===============================
from .utils_ahc import (
    ahc_data, 
    ahc_proses, 
    preprocessing_data, 
    proses_ahc, 
    ahc_hasil, 
    reset_ahc,
    ahc_ai_explain,
    ahc_rekomendasi
)


#profile view
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Profile

@login_required
def profile(request):
    from .models import Polres
    
    user = request.user

    # 🔥 PASTIKAN PROFILE ADA
    profile, created = Profile.objects.get_or_create(user=user)

    if request.method == "POST":
        user.username = request.POST.get('username', '').strip()
        user.email = request.POST.get('email', '').strip()

        full_name = request.POST.get('full_name', '').strip().split(" ")
        user.first_name = full_name[0] if len(full_name) > 0 else ''
        user.last_name = " ".join(full_name[1:]) if len(full_name) > 1 else ''

        user.save()

        profile.alamat = request.POST.get('alamat')
        
        # Assign Polres instance, not string
        polres_id = request.POST.get('polres')
        if polres_id:
            try:
                profile.polres = Polres.objects.get(id=polres_id)
            except Polres.DoesNotExist:
                profile.polres = None
        else:
            profile.polres = None

        if request.FILES.get('foto'):
            profile.foto = request.FILES.get('foto')

        profile.save()
        messages.success(request, 'Profil berhasil diperbarui.')

        return redirect('profile')

    context = {
        'polres_choices': [(p.id, p.nama) for p in Polres.objects.all()]
    }
    return render(request, 'profile.html', context)

# ==========================================
# DATA LAKA MENTAH MANAGEMENT (CLUSTERING RAW DATA)
# ==========================================

@login_required(login_url='login')
def laka_mentah_list(request):
    all_data = LakaMentah.objects.all()
    
    # Pencarian
    search = request.GET.get('search')
    if search:
        all_data = all_data.filter(
            Q(lap_pol__icontains=search) |
            Q(tanggal__icontains=search) |
            Q(tkp__icontains=search) |
            Q(uraian_kejadian__icontains=search)
        )
        
    # Filter Tahun
    selected_tahun = request.GET.get('tahun')
    if selected_tahun:
        all_data = all_data.filter(tahun=selected_tahun)
        
    # Filter Polres
    selected_polres = request.GET.get('polres')
    if selected_polres and selected_polres.isdigit():
        all_data = all_data.filter(polres_id=int(selected_polres))
    
    total_count = all_data.count()
    
    # Hitung duplikasi berdasarkan Nomor Laporan Polisi (LAP. POL) yang tidak kosong
    duplicate_groups = all_data.exclude(lap_pol='').exclude(lap_pol__isnull=True).values(
        'lap_pol'
    ).annotate(count=Count('id')).filter(count__gt=1)
    
    jumlah_duplikat = sum(group['count'] - 1 for group in duplicate_groups)

    # Ambil rincian data duplikat untuk modal
    duplicate_data_details = []
    for group in duplicate_groups:
        example = all_data.filter(lap_pol=group['lap_pol']).first()
        if example:
            duplicate_data_details.append({
                'example': example,
                'count': group['count']
            })

    data_list = all_data.order_by('-id')  # data terbaru di atas

    # Paginasi 20 item per halaman
    paginator = Paginator(data_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Generate list tahun 2027 down to 1945
    years_list = list(range(2027, 1944, -1))
    
    # Get distinct years from database
    filter_years = list(LakaMentah.objects.exclude(tahun__isnull=True).values_list('tahun', flat=True).distinct().order_by('-tahun'))
    
    selected_tahun_int = int(selected_tahun) if selected_tahun and selected_tahun.isdigit() else None
    selected_polres_int = int(selected_polres) if selected_polres and selected_polres.isdigit() else None
    
    context = {
        'data_list': page_obj,
        'total_data': total_count,
        'jumlah_duplikat': jumlah_duplikat,
        'duplicate_data_details': duplicate_data_details,
        'search_query': search or '',
        'years_list': years_list,
        'filter_years': filter_years,
        'selected_tahun': selected_tahun_int,
        'selected_polres': selected_polres_int,
        'polres_list': Polres.objects.all(),
    }
    return render(request, 'coreapp/laka_mentah/list.html', context)


@login_required(login_url='login')
def laka_mentah_tambah(request):
    if request.method == "POST":
        tahun_val = request.POST.get('tahun')
        tahun = int(tahun_val) if tahun_val and tahun_val.isdigit() else None
        polres_id = request.POST.get('polres')
        polres_instance = None
        if polres_id and polres_id.isdigit():
            polres_instance = get_object_or_404(Polres, id=int(polres_id))

        LakaMentah.objects.create(
            tanggal=request.POST.get('tanggal', '').strip(),
            lap_pol=request.POST.get('lap_pol', '').strip(),
            uraian_kejadian=request.POST.get('uraian_kejadian', '').strip(),
            tkp=request.POST.get('tkp', '').strip(),
            terlapor=request.POST.get('terlapor', '').strip(),
            korban=request.POST.get('korban', '').strip(),
            bb=request.POST.get('bb', '').strip(),
            ket=request.POST.get('ket', '').strip(),
            tahun=tahun,
            polres=polres_instance
        )
        messages.success(request, "Data Automatly Report berhasil ditambahkan secara manual.")
        return redirect('laka_mentah_list')

    years_list = list(range(2027, 1944, -1))
    polres_list = Polres.objects.all()
    return render(request, 'coreapp/laka_mentah/tambah.html', {'years_list': years_list, 'polres_list': polres_list})


@login_required(login_url='login')
def laka_mentah_edit(request, pk):
    data = get_object_or_404(LakaMentah, pk=pk)
    
    if request.method == "POST":
        tahun_val = request.POST.get('tahun')
        data.tahun = int(tahun_val) if tahun_val and tahun_val.isdigit() else None
        polres_id = request.POST.get('polres')
        if polres_id and polres_id.isdigit():
            data.polres = get_object_or_404(Polres, id=int(polres_id))
        else:
            data.polres = None

        data.tanggal = request.POST.get('tanggal', '').strip()
        data.lap_pol = request.POST.get('lap_pol', '').strip()
        data.uraian_kejadian = request.POST.get('uraian_kejadian', '').strip()
        data.tkp = request.POST.get('tkp', '').strip()
        data.terlapor = request.POST.get('terlapor', '').strip()
        data.korban = request.POST.get('korban', '').strip()
        data.bb = request.POST.get('bb', '').strip()
        data.ket = request.POST.get('ket', '').strip()
        data.save()
        
        messages.success(request, "Data Automatly Report berhasil diperbarui.")
        return redirect('laka_mentah_list')

    years_list = list(range(2027, 1944, -1))
    polres_list = Polres.objects.all()
    return render(request, 'coreapp/laka_mentah/edit.html', {'data': data, 'years_list': years_list, 'polres_list': polres_list})


@login_required(login_url='login')
def laka_mentah_import(request):
    if request.method == "POST":
        file = request.FILES.get('file')
        tahun_val = request.POST.get('tahun')
        tahun = int(tahun_val) if tahun_val and tahun_val.isdigit() else None
        polres_id = request.POST.get('polres')
        polres_instance = None
        if polres_id and polres_id.isdigit():
            polres_instance = get_object_or_404(Polres, id=int(polres_id))

        if not file:
            messages.error(request, "Silakan pilih file Excel terlebih dahulu.")
            return redirect('laka_mentah_list')
        
        try:
            # 1. Cari baris header secara dinamis (mencari baris yang mengandung 'LAP. POL' atau 'TANGGAL')
            df_raw = pd.read_excel(file, header=None)
            header_row_idx = None
            for idx, row in df_raw.iterrows():
                row_strs = [str(val).strip().upper() for val in row.values if pd.notna(val)]
                # Periksa kecocokan nama kolom utama
                if any('LAP. POL' in s or 'LAP.POL' in s or 'LAP POL' in s or 'LAPORAN POLISI' in s for s in row_strs):
                    header_row_idx = idx
                    break
            
            # Jika ditemukan, baca ulang file dengan baris header tersebut
            if header_row_idx is not None:
                df = pd.read_excel(file, header=header_row_idx)
            else:
                # Fallback default ke baris ke-6 (0-indexed)
                df = pd.read_excel(file, header=6)
            
            df.columns = df.columns.str.strip()
            
            # 2. Mapping kolom agar sesuai dengan database (case-insensitive & clean)
            col_map = {}
            mapped_targets = set()
            for col in df.columns:
                low = str(col).lower()
                target = None
                if 'tanggal' in low and 'tanggal' not in mapped_targets:
                    target = 'tanggal'
                elif ('lap' in low or 'pol' in low) and 'lap_pol' not in mapped_targets:
                    target = 'lap_pol'
                elif ('uraian' in low or 'kejadian' in low) and 'uraian_kejadian' not in mapped_targets:
                    target = 'uraian_kejadian'
                elif ('tkp' in low or 'lokasi' in low) and 'tkp' not in mapped_targets:
                    target = 'tkp'
                elif 'terlapor' in low and 'terlapor' not in mapped_targets:
                    target = 'terlapor'
                elif 'korban' in low and 'korban' not in mapped_targets:
                    target = 'korban'
                elif ('bb' in low or 'bukti' in low) and 'bb' not in mapped_targets:
                    target = 'bb'
                elif 'ket' in low and 'ket' not in mapped_targets:
                    target = 'ket'
                
                if target:
                    col_map[col] = target
                    mapped_targets.add(target)
            
            df = df.rename(columns=col_map)
            
            # 3. Masukkan data literal baris demi baris secara aman
            count = 0
            for _, row in df.iterrows():
                # Helper untuk mendapatkan nilai tunggal secara aman (mengantisipasi Series duplikat)
                def get_row_value(col_name):
                    val = row.get(col_name)
                    if isinstance(val, pd.Series):
                        non_null = val.dropna()
                        val = non_null.iloc[0] if not non_null.empty else None
                    return val

                tanggal_val = get_row_value('tanggal')
                lap_pol_val = get_row_value('lap_pol')
                tkp_val = get_row_value('tkp')

                # Cek baris kosong atau baris penutup halaman excel
                if (tanggal_val is None or pd.isna(tanggal_val) or str(tanggal_val).strip() == '') and \
                   (lap_pol_val is None or pd.isna(lap_pol_val) or str(lap_pol_val).strip() == '') and \
                   (tkp_val is None or pd.isna(tkp_val) or str(tkp_val).strip() == ''):
                    continue
                
                # Helper untuk membersihkan nilai text/nan
                def clean_val(val):
                    if pd.isna(val) or val is None or str(val).strip().upper() in ['NAN', 'NAT', '-']:
                        return ''
                    return str(val).strip()

                LakaMentah.objects.create(
                    tanggal=clean_val(tanggal_val),
                    lap_pol=clean_val(lap_pol_val),
                    uraian_kejadian=clean_val(get_row_value('uraian_kejadian')),
                    tkp=clean_val(tkp_val),
                    terlapor=clean_val(get_row_value('terlapor')),
                    korban=clean_val(get_row_value('korban')),
                    bb=clean_val(get_row_value('bb')),
                    ket=clean_val(get_row_value('ket')),
                    tahun=tahun,
                    polres=polres_instance
                )
                count += 1
            
            messages.success(request, f"Berhasil mengimpor {count} data Automatly Report secara literal.")
        except Exception as e:
            messages.error(request, f"Gagal mengimpor data: {str(e)}")
            import traceback
            traceback.print_exc()
            
        return redirect('laka_mentah_list')
    
    return redirect('laka_mentah_list')


@login_required(login_url='login')
def laka_mentah_hapus(request, pk):
    LakaMentah.objects.filter(pk=pk).delete()
    messages.success(request, "Data Automatly Report berhasil dihapus.")
    return redirect('laka_mentah_list')


@login_required(login_url='login')
def laka_mentah_hapus_semua(request):
    from django.urls import reverse
    
    queryset = LakaMentah.objects.all()
    
    # Filter Tahun
    tahun = request.GET.get('tahun')
    if tahun:
        queryset = queryset.filter(tahun=tahun)
        
    # Filter Polres
    polres_id = request.GET.get('polres')
    if polres_id and polres_id.isdigit():
        queryset = queryset.filter(polres_id=int(polres_id))
        
    count = queryset.count()
    queryset.delete()
    
    if tahun or polres_id:
        messages.success(request, f"Sebanyak {count} data Automatly Report berdasarkan filter yang dipilih berhasil dibersihkan.")
    else:
        messages.success(request, "Seluruh data Automatly Report berhasil dibersihkan.")
        
    url = reverse('laka_mentah_list')
    query_params = []
    if tahun:
        query_params.append(f"tahun={tahun}")
    if polres_id:
        query_params.append(f"polres={polres_id}")
    if query_params:
        url += "?" + "&".join(query_params)
        
    return redirect(url)


@login_required(login_url='login')
def laka_mentah_hapus_duplikat(request):
    if request.method == "POST":
        duplicate_groups = LakaMentah.objects.exclude(lap_pol='').exclude(lap_pol__isnull=True).values(
            'lap_pol'
        ).annotate(count=Count('id')).filter(count__gt=1)
        
        deleted_count = 0
        for group in duplicate_groups:
            ids = list(LakaMentah.objects.filter(lap_pol=group['lap_pol']).values_list('id', flat=True))
            
            # Pertahankan data pertama (ids[0]), hapus selebihnya
            ids_to_delete = ids[1:]
            LakaMentah.objects.filter(id__in=ids_to_delete).delete()
            deleted_count += len(ids_to_delete)
            
        messages.success(request, f"Berhasil membersihkan {deleted_count} data duplikat berdasarkan Nomor Laporan Polisi.")
    return redirect('laka_mentah_list')
